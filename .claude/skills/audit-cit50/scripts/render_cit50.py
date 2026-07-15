# /// script
# requires-python = ">=3.12"
# dependencies = ["pypdf", "cryptography", "openpyxl"]
# ///
"""Fill `3 CIT50 <company>.pdf` (ภ.ง.ด.50 AcroForm) from a client CONTEXT.md.

Deterministic field-fill only. Identity/address/auditor/director/period fields come
from CONTEXT.md (the same data audit-planning already gathers). Tax-computation
numbers (revenue, COGS, net profit, tax base, credits) are NEVER computed by this
script or by an agent guessing — they must be extracted from the client's own
`4 งบการเงิน*.xlsx` (the `ภาษีเงินได้`/`คำนวณภาษี` sheet if present, else the
accounting กำไร(ขาดทุน)สุทธิ line with the assumption of zero tax adjustments — see
references/cit50-field-map.md) and passed in via CLI flags. Fields with no known
source (ISIC code, attachment page counts, bookkeeping fee) are left blank and
reported as warnings.

Usage (from repository root):
    uv run .claude/skills/audit-cit50/scripts/render_cit50.py "PATH/TO/6_ผลจากสกิล/<client>/CONTEXT.md" [options]

Prints JSON: output path and warnings for any field left blank / any assumption made.
"""
import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl
from pypdf import PdfReader, PdfWriter
from pypdf.generic import NameObject, TextStringObject

SKILL = Path(__file__).resolve().parent.parent
TEMPLATE = SKILL / "assets" / "cit50_template.pdf"
AUDITOR_DB = "Database ข้อมูลผู้สอบ.csv.xlsx"
THAI_MONTHS = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
               "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]


def fail(msg: str) -> None:
    print(json.dumps({"ok": False, "error": msg}, ensure_ascii=False))
    sys.exit(1)


def parse_context(path: Path) -> dict:
    profile = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        m = re.match(r"\|\s*([a-z_0-9]+)\s*\|\s*(.*?)\s*\|", line)
        if m:
            profile[m.group(1)] = m.group(2)
    return profile


def clean(v) -> str | None:
    if not v or str(v).strip() in {"—", "-"} or "⚠" in str(v) or "⟨FILL" in str(v):
        return None
    return re.sub(r"\s{2,}", " ", str(v).strip())


def thai_date_parts(ddmmbbbb: str):
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", ddmmbbbb or "")
    if not m:
        return None
    return int(m.group(1)), int(m.group(2)), int(m.group(3))


def split_nid(tax_id: str) -> dict:
    """NID.0-5: fixed-width segments of the 13-digit tax ID (see reference doc)."""
    d = re.sub(r"\D", "", tax_id)
    if len(d) != 13:
        return {}
    widths = [1, 2, 1, 3, 5, 1]
    out, i = {}, 0
    for n, w in enumerate(widths):
        out[f"NID.{n}"] = d[i:i + w]
        i += w
    return out


def spaced_id(digits: str, widths: list[int]) -> str:
    d = re.sub(r"\D", "", digits or "")
    if len(d) != sum(widths):
        return ""
    out, i = [], 0
    for w in widths:
        out.append(d[i:i + w])
        i += w
    return " ".join(out)


def fmt_num(v: float) -> str:
    """2 decimals when meaningful, no trailing zeros/point, never scientific notation."""
    s = f"{v:.2f}".rstrip("0").rstrip(".")
    return s if s else "0"


def split_address(addr: str) -> dict:
    """TXP.6/7/10/11/12/13: เลขที่/หมู่ที่/ถนน/ตำบล/อำเภอ/จังหวัด — best-effort regex parse."""
    out = {}
    m = re.search(r"เลขที่\s*([^\s]+)", addr)
    if not m:
        # some CONTEXT addresses omit the "เลขที่" label and start with the number directly
        m = re.match(r"\s*(\d[\d/]*)", addr)
    if m:
        out["TXP.6"] = m.group(1)
    m = re.search(r"หมู่(?:ที่)?\s*(\d+)", addr)
    if m:
        out["TXP.7"] = m.group(1)
    m = re.search(r"(?:ถนน|ถ\.)\s*([^\s]+(?:\s[^\s]+)?)", addr)
    if m:
        out["TXP.10"] = m.group(1)
    m = re.search(r"ตำบล([^\s]+)|ต\.([^\s]+)", addr)
    if m:
        out["TXP.11"] = m.group(1) or m.group(2)
    m = re.search(r"อำเภอ([^\s]+)|อ\.([^\s]+)", addr)
    if m:
        out["TXP.12"] = m.group(1) or m.group(2)
    m = re.search(r"จังหวัด([^\s]+)|จ\.([^\s]+)", addr)
    if m:
        out["TXP.13"] = m.group(1) or m.group(2)
    return out


def auditor_national_id(first_or_full: str) -> str:
    root = Path.cwd()
    if not (root / AUDITOR_DB).exists():
        return ""
    wb = openpyxl.load_workbook(root / AUDITOR_DB, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[idx.get("ชื่อ", -1)] or "")
        if name and name in first_or_full:
            return spaced_id(str(row[idx.get("เลข บัตร ปชช", -1)] or ""), [1, 4, 5, 2, 1])
    return ""


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("context")
    ap.add_argument("--filing-date", help="dd/mm/bbbb, defaults to CONTEXT sign_date")
    ap.add_argument("--director", help="director/liquidator name (CEO_NAME1)")
    ap.add_argument("--director-title", help="default 'กรรมการ', or 'ผู้ชำระบัญชี' for liquidation (auto-detected)")
    ap.add_argument("--business-desc", help="tightened one-line ประเภทกิจการ; default uses raw CONTEXT business_type")
    ap.add_argument("--isic", help="ISIC code (23) — not in any DB, ask the client's accountant if needed")
    ap.add_argument("--bookkeeping-fee", type=float, help="ค่าทำบัญชี (P5.7.2.13) — not in CONTEXT")
    # Tax computation — extracted by the agent from 4 งบการเงิน*.xlsx, never computed here.
    ap.add_argument("--revenue", type=float, help="P3.2.2.1 (principal) or P3.2.2.4 (other) per --revenue-is-principal")
    ap.add_argument("--revenue-is-principal", action="store_true",
                     help="classify --revenue as P3.2.2.1 (รายได้โดยตรง) instead of the P3.2.2.4 (รายได้อื่น) default")
    ap.add_argument("--cogs", type=float, help="P3.2.2.2 ต้นทุนขาย")
    ap.add_argument("--other-income", type=float, help="P4.5.2.7/.3.7 รายได้อื่น (interest etc.)")
    ap.add_argument("--sga-total", type=float, help="P5.7.2.24/.3.24 รวมรายจ่ายขายและบริหาร")
    ap.add_argument("--net-profit-accounting", type=float, help="กำไรสุทธิทางบัญชี, from the งบการเงิน sheet")
    ap.add_argument("--net-profit-tax", type=float,
                     help="กำไรสุทธิทางภาษีอากร; defaults to --net-profit-accounting (Path B: no adjustments)")
    ap.add_argument("--tax-base", type=float, help="Cit1 ฐานในการคำนวณภาษี (after SME 300k exemption)")
    ap.add_argument("--tax-computed", type=float, help="Cit2 ภาษีที่คำนวณได้")
    ap.add_argument("--wht-credit", type=float, help="Cit5 ภาษีหัก ณ ที่จ่าย")
    ap.add_argument("--pnd51-credit", type=float, help="Cit6 ภาษี ภงด.51 จ่ายล่วงหน้า")
    ap.add_argument("--tax-to-pay", type=float, help="TaxToPay ภาษีที่ชำระเพิ่มเติม; defaults to tax_computed - credits")
    # Balance sheet
    ap.add_argument("--cash", type=float, help="P6.9.1.1.1 เงินสดและรายการเทียบเท่า")
    ap.add_argument("--other-current-assets", type=float, help="P6.9.1.1.4")
    ap.add_argument("--net-fixed-assets", type=float, help="P6.9.2.1.3 ทรัพย์สินหักค่าเสื่อมแล้ว")
    ap.add_argument("--trade-payables", type=float, help="P6.9.3.1.2 เจ้าหนี้การค้า")
    ap.add_argument("--other-current-liabilities", type=float, help="P6.9.3.1.4/P6.9.4.1.3")
    ap.add_argument("--paid-up-capital", type=float, help="P6.9.5.1; defaults to CONTEXT registered_capital")
    ap.add_argument("--retained-earnings", type=float, help="P6.9.5.3 กำไรสะสม/ขาดทุนสะสม")
    ap.add_argument("--total-assets", type=float,
                     help="P6.9.6.2 (grand total assets = grand total liabilities+equity) — same figure as "
                          "audit-planning's --materiality-base, reuse it")
    ap.add_argument("--office-tax-id", help="13-digit เลขประจำตัวผู้เสียภาษีอากร ของสำนักงานสอบบัญชี (aud.3) — "
                                             "varies by which auditor's office signed; not in any DB yet, ask the firm")
    args = ap.parse_args()

    ctx_path = Path(args.context)
    if not ctx_path.is_file():
        fail(f"CONTEXT.md not found: {ctx_path}")
    if not TEMPLATE.exists():
        fail(f"template asset missing: {TEMPLATE} — run build_template.py first")
    prof = parse_context(ctx_path)

    company = clean(prof.get("company_legal_name")) or clean(prof.get("company_name"))
    tax_id = clean(prof.get("tax_id"))
    address = clean(prof.get("address"))
    period_end = clean(prof.get("period_end"))
    period_start_ctx = clean(prof.get("period_start"))
    auditor_name = clean(prof.get("auditor_name"))
    license_no = clean(prof.get("auditor_license"))
    audit_fee = clean(prof.get("audit_fee"))
    registered_capital = clean(prof.get("registered_capital"))
    business_type_text = clean(prof.get("business_type"))
    sign_date = clean(prof.get("sign_date"))

    required = {"company_name": company, "tax_id": tax_id, "period_end": period_end,
                "auditor_name": auditor_name, "auditor_license": license_no}
    missing = [k for k, v in required.items() if not v]
    if missing:
        fail(f"unresolved required field(s) in CONTEXT (⚠ or missing): {missing}. Resolve before rendering.")

    warnings = []

    pe = thai_date_parts(period_end)
    if not pe:
        fail(f"period_end not a dd/mm/bbbb date: {period_end!r}")
    d, mo, y = pe
    ps = thai_date_parts(period_start_ctx) if period_start_ctx else None
    if not ps:
        ps = (1, 1, y)

    status = prof.get("juristic_status", "")
    is_liq = any(k in status for k in ("เลิก", "ชำระบัญชี", "ร้าง")) or "งบเลิก" in ctx_path.parent.name

    filing_date = args.filing_date
    if not filing_date:
        m = re.match(r"(\d{1,2})\s+(\S+)\s+(\d{4})", sign_date or "")
        if m and m.group(2) in THAI_MONTHS:
            filing_date = f"{int(m.group(1)):02d}/{THAI_MONTHS.index(m.group(2)):02d}/{m.group(3)}"
            warnings.append(f"--filing-date not given, assumed CONTEXT sign_date ({sign_date}) — verify against the actual filing date")
        else:
            warnings.append("no --filing-date and CONTEXT sign_date unparseable — fields 46/47/48 left blank")
    fd = thai_date_parts(filing_date) if filing_date else None

    values: dict[str, str] = {}

    values.update(split_nid(tax_id))
    values["TXP.1"] = company
    if address:
        values.update(split_address(address))
    else:
        warnings.append("no address in CONTEXT — TXP.6/7/10/11/12/13 left blank")

    business_desc = clean(args.business_desc) if args.business_desc else business_type_text
    if business_desc:
        values["23.1"] = business_desc
    if args.isic:
        values["24"] = args.isic
    else:
        warnings.append("no --isic given — field 24 (ISIC) left blank, no lookup table exists for this")

    values["17"] = f"{ps[0]:02d}"
    values["18"] = f"{ps[1]:02d}"
    values["19"] = str(ps[2])
    values["20"] = f"{d:02d}"
    values["21"] = f"{mo:02d}"
    values["22"] = str(y)

    if fd:
        values["46"] = f"{fd[0]:02d}"
        values["47"] = f"{fd[1]:02d}"
        values["48"] = str(fd[2])

    values["aud.1"] = f" {auditor_name}"
    values["aud.2"] = license_no.zfill(8)
    nid = auditor_national_id(auditor_name)
    if nid:
        values["aud.0"] = nid
    else:
        warnings.append("auditor's personal tax ID (aud.0) not found in Database ข้อมูลผู้สอบ.csv.xlsx — left blank")

    director = args.director or clean(prof.get("director_1"))
    if director:
        values["CEO_NAME1"] = director
        values["CEO1"] = args.director_title or ("ผู้ชำระบัญชี" if is_liq else "กรรมการ")
    else:
        warnings.append("no director/liquidator name — pass --director NAME (CEO_NAME1/CEO1 left blank)")

    if args.bookkeeping_fee is not None:
        values["P5.7.2.13"] = fmt_num(args.bookkeeping_fee)
        values["P5.7.3.13"] = fmt_num(args.bookkeeping_fee)
    else:
        warnings.append("no --bookkeeping-fee given — P5.7.2.13/.3.13 (ค่าทำบัญชี) left blank")
    if audit_fee:
        fee_num = re.sub(r"[^\d.]", "", audit_fee)
        values["P5.7.2.14"] = fee_num
        values["P5.7.3.14"] = fee_num

    # Tax computation — pure passthrough of agent-supplied numbers, no computation here.
    net_profit_tax = args.net_profit_tax
    if net_profit_tax is None and args.net_profit_accounting is not None:
        net_profit_tax = args.net_profit_accounting
        warnings.append("--net-profit-tax not given, assumed equal to --net-profit-accounting "
                         "(Path B: no tax adjustments) — verify no add-back items apply")

    if args.revenue is not None:
        rev = fmt_num(args.revenue)
        if args.revenue_is_principal:
            values["P3.2.2.1"] = rev
            values["P3.2.3.1"] = rev
        else:
            # Ground truth ([103]): a client whose entire income is "other income" (not
            # its registered principal activity) carries the same figure through the
            # P3.2.2.4/.5/.7 subtotal chain AND into the P4.5 "other income" detail
            # schedule that feeds it — it is one figure appearing in two places on the
            # form, not two different numbers.
            for f in ("P3.2.2.4", "P3.2.2.5", "P3.2.2.7", "P3.2.3.4", "P3.2.3.5", "P3.2.3.7",
                       "P4.5.2.3", "P4.5.2.7", "P4.5.3.3", "P4.5.3.7"):
                values[f] = rev
    else:
        warnings.append("no --revenue given — P3.2.2.1/.4 and the P4.5 other-income schedule left blank")
    if args.cogs is not None:
        values["P3.2.2.2"] = fmt_num(args.cogs)
        values["P3.2.3.2"] = fmt_num(args.cogs)
    if args.other_income is not None:
        # additional other income (e.g. interest) on top of a principal-revenue business
        values["P4.5.2.3"] = fmt_num(args.other_income)
        values["P4.5.2.7"] = fmt_num(args.other_income)
        values["P4.5.3.3"] = fmt_num(args.other_income)
        values["P4.5.3.7"] = fmt_num(args.other_income)
    if args.sga_total is not None:
        values["P3.2.2.8"] = fmt_num(args.sga_total)
        values["P3.2.3.8"] = fmt_num(args.sga_total)
        values["P5.7.2.24"] = fmt_num(args.sga_total)
        values["P5.7.3.24"] = fmt_num(args.sga_total)

    if net_profit_tax is not None:
        npt = fmt_num(net_profit_tax)
        for f in ("P3.2.2.9", "P3.2.2.12", "P3.2.2.14", "P3.2.2.16", "P3.2.2.20",
                   "P3.2.3.9", "P3.2.3.12", "P3.2.3.14", "P3.2.3.16", "P3.2.3.20", "P3.2.3.21"):
            values[f] = npt
        if net_profit_tax < 0:
            values.update({"P3.rdo2": "/2", "P3.rdo3": "/2", "Group91": "/2", "Group5": "/2"})
            warnings.append("net profit is negative — flipped P3.rdo2/P3.rdo3/Group91/Group5 to the loss state (/2)")
    else:
        warnings.append("no --net-profit-tax/--net-profit-accounting given — the whole P3 net-profit cascade left blank")

    if args.tax_base is not None:
        values["Cit1"] = fmt_num(args.tax_base)
    if args.tax_computed is not None:
        values["Cit2"] = fmt_num(args.tax_computed)
    if args.wht_credit is not None:
        values["Cit5"] = fmt_num(args.wht_credit)
    if args.pnd51_credit is not None:
        values["Cit6"] = fmt_num(args.pnd51_credit)

    tax_to_pay = args.tax_to_pay
    if tax_to_pay is None and args.tax_computed is not None:
        tax_to_pay = args.tax_computed - (args.wht_credit or 0) - (args.pnd51_credit or 0)
    if tax_to_pay is not None:
        if tax_to_pay > 0:
            values["TaxToPay"] = fmt_num(tax_to_pay)
            baht, satang = divmod(round(tax_to_pay * 100), 100)
            values["404"] = values["408"] = f" {baht:,}"
            values["405"] = values["409"] = f"{satang:02d}"
        else:
            values["404"] = values["405"] = values["406"] = values["407"] = values["408"] = values["409"] = "-"

    if args.cash is not None:
        values["P6.9.1.1.1"] = fmt_num(args.cash)
    if args.other_current_assets is not None:
        values["P6.9.1.1.4"] = fmt_num(args.other_current_assets)
    if args.net_fixed_assets is not None:
        values["P6.9.2.1.3"] = fmt_num(args.net_fixed_assets)
    if args.trade_payables is not None:
        values["P6.9.3.1.2"] = fmt_num(args.trade_payables)
    if args.other_current_liabilities is not None:
        values["P6.9.3.1.4"] = fmt_num(args.other_current_liabilities)
        values["P6.9.4.1.3"] = fmt_num(args.other_current_liabilities)
    paid_up = args.paid_up_capital
    if paid_up is None and registered_capital:
        paid_up = float(re.sub(r"[^\d.]", "", registered_capital))
    if registered_capital:
        values["P6.9.5.0"] = re.sub(r"[^\d.]", "", registered_capital)
    if paid_up is not None:
        values["P6.9.5.1"] = fmt_num(paid_up)
    if args.retained_earnings is not None:
        values["P6.9.5.3"] = fmt_num(args.retained_earnings)
    if args.total_assets is not None:
        values["P6.9.6.2"] = fmt_num(args.total_assets)
    else:
        warnings.append("no --total-assets given — P6.9.6.2 (grand total, cross-check against Planning's "
                         "materiality_base) left blank")

    if args.office_tax_id:
        values["aud.3"] = spaced_id(args.office_tax_id, [1, 4, 5, 2, 1])
    else:
        warnings.append("no --office-tax-id given — aud.3 (สำนักงานสอบบัญชี tax ID) left blank; "
                         "varies by which auditor's office signed, not in Database ข้อมูลผู้สอบ.csv.xlsx yet")

    # Confirmed constant across all 9 ground-truth cases — see references/cit50-field-map.md
    for f in ("Cit11", "P3.2.1.4", "P4.4.1.6", "P4.4.1.15", "P4.4.1.17", "Text494", "Text496",
              "TotalEX1", "TotalEX2", "TotalEX3",
              "SUM1EX1.19", "SUM2EX1.19", "SUM3EX1.19",
              "SUM1EX2.26", "SUM2EX2.26", "SUM3EX2.26",
              "SUM1EX3.18", "SUM2EX3.18", "SUM3EX3.18"):
        if f not in values:
            values[f] = "0" if not f.startswith("Text") else "-"
    if "404" not in values:
        values.update({"404": "-", "405": "-", "406": "-", "407": "-", "408": "-", "409": "-"})

    if args.cash is None and args.trade_payables is None:
        warnings.append("no balance-sheet figures given (--cash/--trade-payables/...) — P6.9.* section left blank")

    reader = PdfReader(TEMPLATE)
    writer = PdfWriter()
    writer.append(reader)
    # Thai text has no embedded appearance stream for arbitrary values — tell every
    # viewer to regenerate field appearances from the form's own font at display time
    # instead of trying to rasterize it ourselves (which needs fontTools and still
    # risks glyph corruption for a government-supplied Thai font).
    writer.set_need_appearances_writer(True)

    # Many CIT50 fields share a local /T name across different parents (e.g. "17" for
    # day-of-period-start also exists nested as P3.2.1.17, COL1EX1.17, ...).
    # PdfWriter.update_page_form_field_values matches on local /T and would silently
    # cross-write into all of them — so fields are located by FULLY QUALIFIED name via
    # get_fields() and written directly to their own object, never by short name.
    live_fields = writer.get_fields() or {}
    unmatched = []
    for name, value in values.items():
        f = live_fields.get(name)
        if f is None or f.indirect_reference is None:
            unmatched.append(name)
            continue
        obj = writer.get_object(f.indirect_reference)
        obj[NameObject("/V")] = TextStringObject(str(value))
    if unmatched:
        warnings.append(f"field name(s) not found in the template, skipped: {unmatched}")

    out_dir = ctx_path.parent / "WP"
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = f"3 CIT50 {company} {d:02d}{mo:02d}{str(y)[-2:]}.pdf"
    out_path = out_dir / fname
    with open(out_path, "wb") as fh:
        writer.write(fh)

    print(json.dumps({
        "ok": True,
        "output": str(out_path),
        "is_liquidation": is_liq,
        "warnings": warnings,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
