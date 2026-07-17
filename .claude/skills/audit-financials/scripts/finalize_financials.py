# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl"]
# ///
"""Finalize a client's financial-statement working paper (skill 5.5, audit-financials).

Runs AFTER a human has adjusted the scaffold that `audit-workpaper` (skill 5) produced —
i.e. the trial balance is adjusted, every account is classified in `Mapping!H`, and the
adjusting entries are posted. It does two things, and it NEVER freezes formulas or invents
numbers (the workbook stays a live, editable Excel that recomputes itself):

  1. VALIDATION / QA gate — reads the finished workbook's cached values and checks it is
     submission-ready: the balance sheet balances, no error (#REF!/#DIV0) cells remain,
     every balance-bearing account is classified with a valid NPAE caption, the embedded
     company name/year match CONTEXT, and profit ties across the statements. Emits an
     exception report. If a HARD error is found (or the file was never opened/saved in a
     real spreadsheet app, so the formulas have no cached values), it refuses and stops.

  2. NOTE-DETAIL EXPANSION — the scaffold leaves one summary row per caption; the finished
     งบ needs the itemized per-account breakdown under each note. This rewrites the note
     section with, for each caption that has classified accounts, a numbered note header +
     one live row per account (linked to its Mapping row) + a subtotal. The statement lines
     are re-pointed to a self-contained live SUMIF-by-caption so nothing depends on note-row
     positions (openpyxl does not adjust formula refs when rows move). Output is a NEW copy;
     the auditor's working file is never modified in place.

The shared structure this reads is specified in `docs/financials-contract.md`. v1 covers
บจ. going-concern only.

Usage (from repository root):
    uv run .claude/skills/audit-financials/scripts/finalize_financials.py \
        "PATH/TO/6_ผลจากสกิล/<client>/CONTEXT.md" \
        [--wp PATH] [--validate-only] [--out PATH]

Prints JSON: validation report (errors/warnings/checks) + output path + notes_expanded.
"""
import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.workbook.defined_name import DefinedName

FS = "งบการเงิน"
MAP = "Mapping"
TB = "TB"
TAX = "ภาษีเงินได้"
# Print contract (docs/financials-contract.md §10) — must match the scaffold's, or the
# rebuilt note rows print in a different typeface/format from the statements above them.
NUMFMT = r"_(* #,##0.00_);_(* \(#,##0.00\);_(* \-??_);_(@_)"
FS_FONT = "Browallia New"
FS_SIZE = 14
ROW_H = 21.0
NOTE_MARKER = "รายละเอียดประกอบ"       # scaffold's reserved note-detail section header (col A)
TOL = 0.01                              # rounding tolerance for numeric ties (baht)

# ── Controlled caption vocabulary + natural side (must match docs/financials-contract.md §4
#    and audit-workpaper/scripts/scaffold_workpaper.py — keep byte-identical). ────────────
CUR_ASSETS = [
    ("เงินสดและรายการเทียบเท่าเงินสด", "C"),
    ("ลูกหนี้การค้าและลูกหนี้หมุนเวียนอื่น", "C"),
    ("เงินให้กู้ยืมระยะสั้น", "C"),
    ("สินค้าคงเหลือ", "C"),
    ("สินทรัพย์หมุนเวียนอื่น", "C"),
]
NONCUR_ASSETS = [
    ("ที่ดิน อาคารและอุปกรณ์", "C"),
    ("สินทรัพย์ไม่มีตัวตน", "C"),
    ("สินทรัพย์ไม่หมุนเวียนอื่น", "C"),
]
CUR_LIAB = [
    ("เจ้าหนี้การค้าและเจ้าหนี้หมุนเวียนอื่น", "D"),
    ("เงินกู้ยืมระยะสั้น", "D"),
    ("ภาษีเงินได้ค้างจ่าย", "D"),
]
NONCUR_LIAB = [
    ("เงินกู้ยืมระยะยาว", "D"),
    ("หนี้สินไม่หมุนเวียนอื่น", "D"),
]
SHARE_CAP = ("ทุนที่ออกและชำระแล้ว", "D")
RETAINED = ("กำไร(ขาดทุน)สะสม", "D")
REVENUE = [("รายได้จากการขายหรือบริการ", "D"), ("รายได้อื่น", "D")]
EXPENSES = [
    ("ต้นทุนขายหรือบริการ", "C"),
    ("ค่าใช้จ่ายในการขายและบริหาร", "C"),
    ("ต้นทุนทางการเงิน", "C"),
]
ALL_CAPTIONS = (
    [c for c, _ in CUR_ASSETS + NONCUR_ASSETS]
    + [c for c, _ in CUR_LIAB + NONCUR_LIAB]
    + [SHARE_CAP[0], RETAINED[0]]
    + [c for c, _ in REVENUE + EXPENSES]
)
SIDE = dict(CUR_ASSETS + NONCUR_ASSETS + CUR_LIAB + NONCUR_LIAB + [SHARE_CAP, RETAINED] + REVENUE + EXPENSES)
CAPSET = set(ALL_CAPTIONS)

def F(**kw) -> Font:
    """A font in the locked FS typeface (§10.1). Always name it: a bare Font(bold=True)
    carries no name and Excel silently falls back to Calibri."""
    return Font(name=FS_FONT, size=FS_SIZE, **kw)


PLAIN = F()
BOLD = F(bold=True)
HEAD = F(bold=True)
WARN = Font(color="9C6210")
THIN = Side(style="thin", color="BBBBBB")
TOPBORDER = Border(top=THIN)


# ── helpers ─────────────────────────────────────────────────────────────────────────
def fail(msg: str) -> None:
    print(json.dumps({"ok": False, "error": msg}, ensure_ascii=False))
    sys.exit(1)


def parse_context(path: Path) -> dict:
    profile = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        m = re.match(r"\|\s*([a-z_0-9]+)\s*\|\s*(.*?)\s*\|", line)
        if m:
            profile[m.group(1)] = m.group(2)
    return profile


def clean(v):
    if not v or str(v).strip() in {"—", "-"} or "⚠" in str(v) or "⟨FILL" in str(v):
        return None
    return re.sub(r"\s{2,}", " ", str(v).strip())


def be_year(period: str):
    m = re.search(r"/(\d{4})\s*$", period or "")
    return int(m.group(1)) if m else None


def norm(s):
    """Normalize a company name for tolerant comparison (drop spaces/entity words)."""
    if s is None:
        return ""
    s = re.sub(r"\s+", "", str(s))
    for w in ("บริษัท", "จำกัด", "ห้างหุ้นส่วนจำกัด", "ห้างหุ้นส่วนสามัญ", "(มหาชน)"):
        s = s.replace(w, "")
    return s


def is_error(v) -> bool:
    """True if a cached cell value is an Excel error literal (#REF!, #DIV/0!, …)."""
    return isinstance(v, str) and v.startswith("#") and v.endswith(("!", "?"))


def num(v):
    return v if isinstance(v, (int, float)) else None


def dn_target(wb, name):
    dn = wb.defined_names.get(name)
    if dn is None:
        return None
    for sheet, coord in dn.destinations:
        return sheet, coord.replace("$", "")
    return None


def dn_value(wb_data, name):
    t = dn_target(wb_data, name)
    if t is None:
        return None
    sheet, coord = t
    try:
        return wb_data[sheet][coord].value
    except Exception:  # noqa: BLE001
        return None


def locate_wp(ctx_path: Path, explicit: str | None) -> Path:
    if explicit:
        p = Path(explicit)
        if not p.exists():
            fail(f"--wp not found: {p}")
        return p
    wp_dir = ctx_path.parent / "WP"
    cands = sorted(wp_dir.glob("4 งบการเงิน*.xlsx")) if wp_dir.exists() else []
    cands = [c for c in cands if not c.name.startswith("~$")]
    if not cands:
        fail(f"no '4 งบการเงิน*.xlsx' found under {wp_dir}; run audit-workpaper (skill 5) first, "
             "or pass --wp PATH")
    return cands[-1]


# ── validation (reads cached values; never mutates) ──────────────────────────────────
def validate(wb_data, wb_form, profile: dict) -> dict:
    errors: list[str] = []
    warnings: list[str] = []
    checks: dict = {}

    company_ctx = clean(profile.get("company_legal_name")) or clean(profile.get("company_name"))
    cy = be_year(profile.get("period_end", ""))
    py = be_year(profile.get("prior_period_end", "")) or (cy - 1 if cy else None)

    # (1) recalc guard: FS total cells hold formulas; if data_only reads None the workbook was
    #     never opened & saved by a real spreadsheet app, so there are no cached values to read.
    ta = dn_value(wb_data, "FS_TOTAL_ASSETS_CY")
    tle = dn_value(wb_data, "FS_TOTAL_LIAB_EQUITY_CY")
    if FS not in wb_form.sheetnames:
        errors.append(f"sheet '{FS}' not found — is this a งบการเงิน workpaper from skill 5?")
        return {"errors": errors, "warnings": warnings, "checks": checks}
    if ta is None and tle is None:
        errors.append(
            "the workbook has no calculated values — openpyxl cannot evaluate formulas, so this "
            "file must be opened once in Excel/Google Sheets/LibreOffice and SAVED (which caches "
            "the computed values) before it can be validated. Open it, save, and re-run.")
        return {"errors": errors, "warnings": warnings, "checks": checks, "unsaved": True}

    # (2) error cells anywhere (#REF!, #DIV/0!, broken external links → refuse)
    err_cells = []
    for ws in wb_data.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if is_error(c.value):
                    err_cells.append(f"{ws.title}!{c.coordinate}={c.value}")
    if err_cells:
        errors.append(f"{len(err_cells)} error cell(s) present — refuse to finalize a งบ with errors: "
                      + ", ".join(err_cells[:8]) + (" …" if len(err_cells) > 8 else ""))
    checks["error_cells"] = len(err_cells)

    # (3) balance sheet balances (CY and PY)
    for tag, tot_a, tot_le in (("CY", "FS_TOTAL_ASSETS_CY", "FS_TOTAL_LIAB_EQUITY_CY"),
                               ("PY", "FS_TOTAL_ASSETS_PY", "FS_TOTAL_LIAB_EQUITY_PY")):
        a, le = num(dn_value(wb_data, tot_a)), num(dn_value(wb_data, tot_le))
        if a is None or le is None:
            warnings.append(f"could not read {tag} balance-sheet totals to check balancing")
            continue
        diff = round(a - le, 2)
        checks[f"bs_diff_{tag.lower()}"] = diff
        if abs(diff) > TOL:
            errors.append(f"balance sheet does NOT balance ({tag}): รวมสินทรัพย์ {a:,.2f} − "
                          f"รวมหนี้สินและส่วนของผู้ถือหุ้น {le:,.2f} = {diff:,.2f} (ควรเป็น 0)")

    # (4) profit tie: IS profit-before-tax must equal the tax sheet's starting profit
    pbt = num(dn_value(wb_data, "TAX_NET_PROFIT"))
    net = num(dn_value(wb_data, "FS_NET_PROFIT_CY"))
    checks["pbt_cy"] = pbt
    checks["net_profit_cy"] = net
    tax_c4 = None
    if TAX in wb_data.sheetnames:
        tax_c4 = num(wb_data[TAX]["C4"].value)   # tax sheet: กำไรสุทธิทางบัญชีก่อนภาษี
    if pbt is not None and tax_c4 is not None and abs(round(pbt - tax_c4, 2)) > TOL:
        errors.append(f"กำไรก่อนภาษีในงบกำไรขาดทุน ({pbt:,.2f}) ไม่ตรงกับยอดตั้งต้นในชีต {TAX} "
                      f"({tax_c4:,.2f})")
    warnings.append("การกระทบยอดกำไรสุทธิกับ ภ.ง.ด.50 (CIT50) ต้องตรวจด้วยมือ — เทียบ "
                    f"FS_NET_PROFIT_CY ({net:,.2f} บาท) กับกำไรสุทธิในแบบ ภ.ง.ด.50 หน้า 3"
                    if net is not None else
                    "ตรวจกระทบยอดกำไรสุทธิกับ ภ.ง.ด.50 (CIT50) ด้วยมือ")

    # (5) built-in tie-out row (=รวมสินทรัพย์ − รวมหนี้สินและส่วนของผู้ถือหุ้น) reads 0
    checkrow = None
    for row in wb_data[FS].iter_rows(min_col=1, max_col=1):
        if row[0].value and "ตรวจสอบ" in str(row[0].value) and "0" in str(row[0].value):
            checkrow = wb_data[FS].cell(row[0].row, 5).value
            break
    if isinstance(checkrow, (int, float)) and abs(checkrow) > TOL:
        errors.append(f"แถวตรวจสอบสมดุลในงบแสดงฐานะการเงินอ่านได้ {checkrow:,.2f} (ต้องเป็น 0)")

    # (6) Mapping!H completeness + valid vocabulary (balance-bearing accounts must be classified)
    mp = wb_form[MAP]
    unclassified, badcaption = [], []
    for r in range(6, mp.max_row + 1):
        code = mp.cell(r, 1).value
        if code in (None, ""):
            continue
        cap = mp.cell(r, 8).value
        # a balance-bearing account is one with any non-zero cached amount in C or D
        cd = num(wb_data[MAP].cell(r, 3).value) or 0
        cc = num(wb_data[MAP].cell(r, 4).value) or 0
        bearing = abs(cd) > TOL or abs(cc) > TOL
        if cap in (None, ""):
            if bearing:
                unclassified.append(f"{code} {mp.cell(r, 2).value or ''}".strip())
        elif str(cap).strip() not in CAPSET:
            badcaption.append(f"{code}→'{cap}'")
    checks["unclassified_accounts"] = len(unclassified)
    if unclassified:
        errors.append(f"{len(unclassified)} บัญชีที่มียอดยังไม่ได้จัดประเภทใน Mapping!H: "
                      + ", ".join(unclassified[:8]) + (" …" if len(unclassified) > 8 else ""))
    if badcaption:
        errors.append(f"{len(badcaption)} บัญชีจัดประเภทเป็น caption นอกรายการมาตรฐาน NPAE: "
                      + ", ".join(badcaption[:8]) + (" …" if len(badcaption) > 8 else ""))

    # (7) embedded company name / year sanity (stale-sheet guard, contract §8)
    if company_ctx:
        for sh in (FS, TB, MAP, TAX):
            if sh not in wb_form.sheetnames:
                continue
            a1 = wb_form[sh]["A1"].value
            if a1 and norm(a1) and norm(company_ctx) not in norm(a1) and norm(a1) not in norm(company_ctx):
                warnings.append(f"ชื่อบริษัทในชีต '{sh}' (A1='{a1}') ไม่ตรงกับ CONTEXT "
                                f"('{company_ctx}') — อาจเป็นชีตค้างจากงานอื่น ตรวจสอบ")
    if cy:
        stale_years = set()
        for sh in wb_form.sheetnames:
            for row in wb_form[sh].iter_rows(max_row=4, max_col=8):
                for c in row:
                    for yy in re.findall(r"25\d\d", str(c.value or "")):
                        if int(yy) not in (cy, py):
                            stale_years.add(yy)
        if stale_years:
            warnings.append(f"พบเลขปีอื่นที่ไม่ใช่ {py}/{cy} ในหัวชีต: {', '.join(sorted(stale_years))} "
                            "— ตรวจว่าไม่ใช่ป้ายปีค้าง")

    return {"errors": errors, "warnings": warnings, "checks": checks}


# ── note-detail expansion (rewrites the note section; keeps everything live) ──────────
def group_mapping_accounts(wb_form) -> dict[str, list[tuple[int, str]]]:
    """caption -> list of (mapping_row, account_name) for accounts classified to it, in order."""
    mp = wb_form[MAP]
    groups: dict[str, list[tuple[int, str]]] = {}
    for r in range(6, mp.max_row + 1):
        code = mp.cell(r, 1).value
        cap = mp.cell(r, 8).value
        if code in (None, "") or cap in (None, ""):
            continue
        cap = str(cap).strip()
        if cap not in CAPSET:
            continue
        name = mp.cell(r, 2).value or str(code)
        groups.setdefault(cap, []).append((r, str(name)))
    return groups


def repaginate(ws) -> None:
    """Re-extend the print area over the rebuilt note rows (contract §10.5).

    The งบ is a printed deliverable whose print area is pinned to `$A$1:$G$<last row>`.
    Expanding the notes changes the last row, so without this the added note rows fall
    outside the printed page. The statement pages' explicit breaks sit above the note
    marker and are deliberately left untouched.
    """
    ws.print_area = f"$A$1:$G${ws.max_row}"
    for r in range(1, ws.max_row + 1):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = ROW_H


def expand_notes(wb_form, groups: dict) -> int:
    """Rewrite the งบการเงิน note-detail section with itemized per-account rows.

    Returns the number of captions expanded. Statement lines are re-pointed to a
    self-contained SUMIF-by-caption so they never depend on note-row positions.
    """
    ws = wb_form[FS]

    # find the note-detail section start (scaffold marker in col A) and drop it + everything after
    note_start = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        if row[0].value and NOTE_MARKER in str(row[0].value):
            note_start = row[0].row
            break
    if note_start is None:
        note_start = ws.max_row + 2  # no reserved section (defensive) — append fresh below content

    # (a) re-point every statement caption line to a self-contained live SUMIF-by-caption,
    #     so deleting/rebuilding the note rows below cannot break them. re_bs (retained earnings
    #     on the BS, which references the equity statement) is left untouched.
    note_no: dict[str, int] = {}
    n = 4  # notes 1-3 are general/basis/policy; caption notes start at 4
    for cap in ALL_CAPTIONS:
        if cap in groups:
            note_no[cap] = n
            n += 1
    for r in range(1, note_start):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        cap = str(a).strip()
        if cap not in CAPSET:
            continue
        # cite the note number on every statement line for this caption (incl. retained
        # earnings on the BS, whose amount comes from the equity statement, not a note ref)
        ws.cell(r, 3).value = note_no.get(cap, "")
        e = ws.cell(r, 5).value
        # rewrite only the lines that pointed into the (old) note area — i.e. '=+E<n>' with
        # n >= note_start. This excludes re_bs ('=+E58', an equity-statement ref above note_start).
        if not (isinstance(e, str) and re.match(r"^=\+E\d+$", e)):
            continue
        ref = int(re.match(r"^=\+E(\d+)$", e).group(1))
        if ref < note_start:
            continue
        pcol = "E" if SIDE[cap] == "C" else "F"
        scol = SIDE[cap]
        ws.cell(r, 5).value = f'=SUMIF({MAP}!$H:$H,"{cap}",{MAP}!${scol}:${scol})'
        ws.cell(r, 7).value = f'=SUMIF({MAP}!$H:$H,"{cap}",{MAP}!${pcol}:${pcol})'

    # (b) delete the old note-detail rows (from the marker to the end of the sheet). Nothing above
    #     now references them (defined names live at rows above note_start; see contract §3.3).
    if note_start <= ws.max_row:
        ws.delete_rows(note_start, ws.max_row - note_start + 1)

    # (c) rebuild: for each caption with accounts, a numbered note header + one row per account
    #     (live-linked to its Mapping row) + a subtotal. All live formulas, fully editable.
    row = note_start
    ws.cell(row, 1, "รายละเอียดประกอบรายการในงบการเงิน (หมายเหตุ 4 เป็นต้นไป)").font = HEAD
    ws.cell(row, 5, "(หน่วย: บาท)").font = F(italic=True)
    row += 2
    expanded = 0
    for cap in ALL_CAPTIONS:
        accts = groups.get(cap)
        if not accts:
            continue
        side = SIDE[cap]
        cy_col = "C" if side == "C" else "D"     # Mapping current-year column
        py_col = "E" if side == "C" else "F"     # Mapping prior-year column
        hdr = ws.cell(row, 1, f"{note_no[cap]}. {cap}")
        hdr.font = BOLD
        row += 1
        first = row
        for maprow, name in accts:
            nm = ws.cell(row, 1, name)
            nm.alignment = Alignment(indent=1)
            nm.font = PLAIN
            ec = ws.cell(row, 5, f"=+{MAP}!{cy_col}{maprow}")
            gc = ws.cell(row, 7, f"=+{MAP}!{py_col}{maprow}")
            ec.number_format = gc.number_format = NUMFMT
            ec.font = gc.font = PLAIN
            row += 1
        st = ws.cell(row, 1, f"รวม{cap}")
        st.font = BOLD
        se = ws.cell(row, 5, f"=SUM(E{first}:E{row - 1})")
        sg = ws.cell(row, 7, f"=SUM(G{first}:G{row - 1})")
        se.font = sg.font = BOLD
        se.border = sg.border = TOPBORDER
        se.number_format = sg.number_format = NUMFMT
        row += 2
        expanded += 1
    return expanded


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("context", help="path to the client's CONTEXT.md")
    ap.add_argument("--wp", help="path to the WP workbook (default: newest '4 งบการเงิน*.xlsx' in <ctx>/WP)")
    ap.add_argument("--validate-only", action="store_true", help="run the QA gate only; do not write a copy")
    ap.add_argument("--out", help="output path for the finalized copy (default: '<wp stem> (final).xlsx')")
    args = ap.parse_args()

    ctx_path = Path(args.context)
    if not ctx_path.exists():
        fail(f"CONTEXT.md not found: {ctx_path}")
    profile = parse_context(ctx_path)
    wp_path = locate_wp(ctx_path, args.wp)

    try:
        wb_data = openpyxl.load_workbook(wp_path, data_only=True)
        wb_form = openpyxl.load_workbook(wp_path, data_only=False)
    except Exception as e:  # noqa: BLE001
        fail(f"could not open WP workbook '{wp_path}': {e}")

    report = validate(wb_data, wb_form, profile)
    hard = bool(report["errors"])

    if args.validate_only or hard:
        out = {"ok": not hard, "mode": "validate-only" if args.validate_only else "validate",
               "wp": str(wp_path), "validation": report, "notes_expanded": 0}
        if hard and not args.validate_only:
            out["stopped"] = "validation failed — fix the exceptions above, re-save in Excel, then re-run"
        print(json.dumps(out, ensure_ascii=False))
        sys.exit(0 if not hard else 1)

    # validation passed → expand note detail into a NEW copy (never touch the working file)
    groups = group_mapping_accounts(wb_form)
    expanded = expand_notes(wb_form, groups)
    repaginate(wb_form[FS])

    out_path = Path(args.out) if args.out else wp_path.with_name(f"{wp_path.stem} (final){wp_path.suffix}")
    wb_form.save(out_path)

    warnings = list(report["warnings"])
    warnings.append("ไฟล์ผลลัพธ์ยังเป็น Excel ที่แก้ไขได้และเป็น formula สด — เปิดใน Excel หนึ่งครั้งเพื่อให้"
                    "คำนวณค่าใหม่ (หมายเหตุรายบัญชีที่เพิ่มจะแสดงยอดหลังเปิด)")
    print(json.dumps({"ok": True, "mode": "finalize", "wp": str(wp_path),
                      "output": str(out_path), "notes_expanded": expanded,
                      "validation": {"errors": [], "warnings": warnings, "checks": report["checks"]}},
                     ensure_ascii=False))


if __name__ == "__main__":
    main()
