# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl"]
# ///
"""Render `1 Planning <company>.xlsx` from a client CONTEXT.md.

Deterministic: fills ONLY sheet `ข้อมูลลูกค้า` (plus the boilerplate risk/TB rows in
`301` and `203 TB (2)`) from a locked client-neutral template. Every other sheet and
cell — including the 9 sheets that cascade off `ข้อมูลลูกค้า` via existing formulas —
is left byte-for-byte untouched. No free-form narrative is generated; sheets 301/203
select one of 3 known boilerplate variants (see references/planning-structure.md).

Usage (from repository root):
    uv run .claude/skills/audit-planning/scripts/render_planning.py "PATH/TO/6_ผลจากสกิล/<client>/CONTEXT.md" [options]

Prints JSON: output path, chosen business_type, and warnings for any field that
could not be resolved from CONTEXT/DB and needs a human (never invented).
"""
import argparse
import datetime
import json
import re
import sys
from pathlib import Path

import openpyxl

SKILL = Path(__file__).resolve().parent.parent
TEMPLATE = SKILL / "assets" / "planning_template.xlsx"
THAI_MONTHS = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
               "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]

BOILERPLATE = {
    "default": {
        "301_row28": ["มีผลกระทบเนื่องจากมีปัญหาเรื่อง อย.", "สินค้าคงเหลือเสื่อมสภาพ ", "V",
                       "มีโอกาสเกิดขึ้น แต่ไม่มาก เนื่องจากสินค้าส่วนใหญ่เป็นสารตั้งต้น และเหลือจำนวนไม่มากนัก", "ไม่"],
        "tb_rows": {
            5: ("รายได้จากการขาย ลดลง", "เนื่องจากสินค้าต้องมีการแก้ไข อย.", None, None),
            6: ("ค่าใช้จ่ายส่วนที่เกี่ยวข้องกับการขาย ก็ลดลงด้วย  เช่น ค่าธรรมเนียม Lazada / ค่าน้ำมัน ", None, None, None),
            7: ("ทั้งนี้มีค่าโฆษณาที่เพิ่มขึ้น เมื่อตรวจทานเอกสาร และสอบถามผู้บริหาร พบว่าเป็นค่า", None, None, None),
            8: ("จัดทำ วิดิโอ โปรโมท ซึ่งเตรียมไว้ใช้โฆษณาหลังจากที่ดำเนินการเรื่อง อย.ให้ถูกต้องเรียบร้อย ", None, None, None),
            10: ("มีเงินกู้ยืมระยะสั้นเพิ่มขึ้น เนื่องจากต้องนำมาใช้หมุนเวียนในการดำเนินการกับกิจการ", None, None, None),
        },
        "materiality_reason": "กิจการมีการรายได้จากการขาย",
        "prior_adjustment_note": "มีรายการปรับปรุงไม่มาก",
        "biz_environment": "กิจการเริ่มจำหน่ายสินค้าได้ ",
    },
    "construction": {
        "301_row28": ["การรับรู้รายได้ตามอัตราร้อยละของงานที่ทำเสร็จ \nอาจทำไม่ได้จริง เนื่องจาก มีการประเมินค่อนข้างยาก อีกทั้งเมื่อพิจารณษจากการบันทึกบัญชี พบว่ากิจการน่าจะบันทึกรายได้ เมื่อได้รับชำระ",
                       "บันทึกรายได้ไม่ครบถ้วน \nบันทึกค่าใช้จ่ายสูงกว่าความเป็นจริง", "C\nA\nCO",
                       "มีโอกาสเกิดขึ้นสูง แต่ไม่ทราบจำนวน ว่ากระทบกับงบการเงินมากน้อยเพียงใด", "ใช่"],
        "tb_rows": {
            5: ("มีรายได้จากการบริการเพิ่มขึ้น", None, "เนื่องจากมีการดำเนินงานเป็นปีแรก", None),
            6: ("ค่าใช้จ่ายส่วนที่เกี่ยวข้อง ทั้งต้นทุน และค่าใช้จ่ายบริหารก็มีจำนวนเพิ่มขึ้นเช่นกัน", None, None, None),
        },
        "materiality_reason": "กิจการมีการดำเนินงานปกติ",
        "prior_adjustment_note": "ปีก่อนมีรายการปรับปรุงไม่มาก",
        "biz_environment": "ไม่มี",
    },
    "dormant": {
        "301_row28": None,
        "tb_rows": {5: ("กิจการไม่มีการดำเนินงาน", None, None, None)},
        "materiality_reason": "กิจการไม่มีรายได้",
        "prior_adjustment_note": "ปีก่อนมีรายการปรับปรุงไม่มาก",
        "biz_environment": "ห้างเลิกกิจการ",
    },
}


def fail(msg: str) -> None:
    print(json.dumps({"ok": False, "error": msg}, ensure_ascii=False))
    sys.exit(1)


def parse_context(path: Path) -> dict:
    profile = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        m = re.match(r"\|\s*([a-z_0-9]+)\s*\|\s*(.*?)\s*\|", line)
        if m:
            profile[m.group(1)] = m.group(2)
    return profile


def clean(v) -> str | None:
    if not v or str(v).strip() in {"—", "-"} or "⚠" in str(v) or "⟨FILL" in str(v):
        return None
    return re.sub(r"\s{2,}", " ", str(v).strip())


def thai_date_parts(ddmmbbbb: str):
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", ddmmbbbb or "")
    if not m:
        return None
    return int(m.group(1)), int(m.group(2)), int(m.group(3))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("context")
    ap.add_argument("--business-type", choices=["default", "construction", "dormant"],
                     help="selects the 301/203 boilerplate variant (see references/planning-structure.md); "
                          "auto-detected as dormant for liquidation, else default")
    ap.add_argument("--director", action="append", default=[], help="director name, repeatable up to 3")
    ap.add_argument("--director-authority", help="e.g. 'กรรมการคนใดคนหนึ่งลงนาม'")
    ap.add_argument("--audit-team", action="append", default=[], help="assistant auditor name, repeatable")
    ap.add_argument("--audit-start-date", help="dd/mm/bbbb")
    ap.add_argument("--audit-hours", type=int)
    ap.add_argument("--audit-expense", type=int, help="override the auto-estimate (30%% of audit_fee)")
    ap.add_argument("--rep-letter-date", help="dd/mm/bbbb, management representation letter date")
    ap.add_argument("--approval-date", help="dd/mm/bbbb, rare field — leave unset unless supplied")
    ap.add_argument("--materiality-basis", choices=["สินทรัพย์รวม", "รายได้รวม"],
                     help="no default — ground truth has no reliable majority (roughly 1:2 "
                          "สินทรัพย์รวม:รายได้รวม, ~19%% left blank outright); ⚠ if omitted")
    ap.add_argument("--materiality-reason", help="overrides the business_type default for row 47")
    ap.add_argument("--materiality-base", type=float, help="numeric base amount, from งบการเงิน")
    ap.add_argument("--prior-adjustment-note", help="overrides the business_type default for row 54")
    ap.add_argument("--period-start", help="dd/mm/bbbb, explicit override for row 9 (default: incorporation_date "
                                             "for a first-year client whose incorporation falls inside the audited "
                                             "period, else 1 Jan of the BE year)")
    ap.add_argument("--period-end", help="dd/mm/bbbb, overrides CONTEXT's period_end for row 10/48 — "
                                          "needed for งบเลิก where the actual dissolution cutoff differs from DB งาน's nominal fiscal year end")
    ap.add_argument("--business-desc", help="tightened one-line ประเภทธุรกิจ phrase; default uses the raw DBD business_type text verbatim")
    ap.add_argument("--biz-environment", help="overrides the business_type default for row 56 (สภาพแวดล้อมของกิจการ, cascades into 301!B20)")
    ap.add_argument("--tb-note", help="overrides the boilerplate 203 TB (2)!B5 text with a client-specific one-liner "
                                       "(use when the client doesn't fit default/construction/dormant cleanly)")
    args = ap.parse_args()

    ctx_path = Path(args.context)
    if not ctx_path.is_file():
        fail(f"CONTEXT.md not found: {ctx_path}")
    if not TEMPLATE.exists():
        fail(f"template asset missing: {TEMPLATE} — run build_template.py first")
    prof = parse_context(ctx_path)

    company = clean(prof.get("company_legal_name")) or clean(prof.get("company_name"))
    entity = prof.get("entity_type", "")
    tax_id = clean(prof.get("tax_id"))
    period_end = args.period_end or clean(prof.get("period_end"))
    prior_period_end = clean(prof.get("prior_period_end"))
    auditor_name = clean(prof.get("auditor_name"))
    license_no = clean(prof.get("auditor_license"))
    auditor_type = prof.get("auditor_type", "")
    audit_fee = clean(prof.get("audit_fee"))
    business_type_text = clean(prof.get("business_type"))

    required = {"company_name": company, "tax_id": tax_id, "period_end": period_end,
                "auditor_name": auditor_name, "auditor_license": license_no, "audit_fee": audit_fee}
    missing = [k for k, v in required.items() if not v]
    if missing:
        fail(f"unresolved required field(s) in CONTEXT (⚠ or missing): {missing}. Resolve before rendering.")

    warnings = []

    # business_type: explicit flag wins; else auto-detect dormant from liquidation signals
    # (a real, legally-grounded signal). Anything else is a judgment call with NO reliable
    # auto-signal — confirmed empirically by cross-checking all 16 ground-truth files:
    # business_type/revenue/juristic_status text does not correlate with which variant the
    # firm actually picked (e.g. a client literally registered as "การก่อสร้างอาคาร..." got
    # `default`, while a landscaping-maintenance client got `construction`). Guessing here
    # would silently ship wrong boilerplate to the client, so refuse instead of defaulting.
    status = prof.get("juristic_status", "")
    is_liq = any(k in status for k in ("เลิก", "ชำระบัญชี", "ร้าง")) or "งบเลิก" in ctx_path.parent.name
    if args.business_type:
        business_type = args.business_type
    elif is_liq:
        business_type = "dormant"
    else:
        fail("--business-type not given and this client is not a detected liquidation case — "
             "'default' vs 'construction' vs 'dormant' is a judgment call with no reliable "
             "auto-signal (business_type/revenue/juristic_status text does not predict it, "
             "confirmed against ground truth). ASK THE USER which variant fits this client "
             "(show them CONTEXT's business_type text and revenue as context), then re-run "
             "with --business-type default|construction|dormant.")
    boiler = BOILERPLATE[business_type]

    pe = thai_date_parts(period_end)
    if not pe:
        fail(f"period_end not a dd/mm/bbbb date: {period_end!r}")
    d, mo, y = pe
    fiscal_year = str(y)

    def be_date(ddmmbbbb: str):
        """dd/mm/bbbb in Thai Buddhist Era (as used by DB งาน and all CLI date flags) -> date."""
        p = thai_date_parts(ddmmbbbb)
        if not p:
            return None
        d, mo, y = p
        return datetime.date(y - 543, mo, d)

    def ce_date(ddmmbbbb: str):
        """dd/mm/bbbb already in Gregorian CE (as returned by the DBD lookup) -> date."""
        p = thai_date_parts(ddmmbbbb)
        if not p:
            return None
        d, mo, y = p
        return datetime.date(y, mo, d)

    incorp = clean(prof.get("incorporation_date"))
    incorp_date = ce_date(incorp) if incorp else None
    # first-year: incorporation falls after the notional 1-Jan period start, i.e. inside the
    # audited period itself — a genuine first-year engagement, not a rollover client.
    is_first_year = incorp_date is not None and incorp_date > datetime.date(y - 543, 1, 1)

    if args.period_start:
        period_start_date = be_date(args.period_start)
    elif is_first_year:
        period_start_date = incorp_date
        warnings.append(f"first-year client detected (incorporation_date {incorp} falls inside the audited "
                         "period) — row 9 (วันต้นงวด) set to incorporation_date instead of the 1-Jan default; "
                         "prior-period rows (11/12) left blank (no prior year existed)")
    else:
        period_start_date = be_date(f"01/01/{y}")

    prior_start = None
    if prior_period_end:
        pp = thai_date_parts(prior_period_end)
        if pp:
            prior_start = f"01/01/{pp[2]}"

    is_company = not entity.startswith("หจก")
    entity_label = "ห้างหุ้นส่วนจำกัด" if not is_company else ("บริษัทจำกัด" if is_liq else "บริษัท")

    def expand_address(addr: str) -> str:
        """Match the firm's own formatting: 'เลขที่ ' prefix (plain space, no line break) + spelled-out
        ตำบล/อำเภอ/จังหวัด. Ground truth across all 16 files: 11/16 add the 'เลขที่ ' prefix (8 with a
        plain space, 3 with a stray newline — space is the majority), 5/16 have no prefix at all with no
        reliable signal in CONTEXT to predict which; space-prefixed is the closest single convention."""
        addr = re.sub(r"\bต\.", "ตำบล", addr)
        addr = re.sub(r"\bอ\.", "อำเภอ", addr)
        addr = re.sub(r"\bจ\.", "จังหวัด", addr)
        return f"เลขที่ {addr}" if not addr.startswith("เลขที่") else addr

    directors = (args.director or [])[:3]
    if not directors:
        d1 = clean(prof.get("director_1"))
        if d1:
            directors = [d1]
    if not directors:
        warnings.append("no director name resolved — pass --director NAME (row 13 กรรมการ1 left blank)")

    team = args.audit_team or []
    if not team:
        warnings.append("no audit team members given — pass --audit-team NAME (repeatable); "
                         "row 20/21 (จำนวนผู้สอบบัญชีในทีม / ชื่อผู้สอบบัญชีในทีม) left blank")

    if not args.audit_start_date:
        warnings.append("no --audit-start-date given — row 26 (ตรวจสอบตั้งแต่วันที่) left blank")
    if args.audit_hours is None:
        warnings.append("no --audit-hours given — row 29 (จำนวนชั่วโมง) left blank; no DB source for this field")
    if not args.rep_letter_date:
        warnings.append("no --rep-letter-date given — row 39 (วันที่ในหนังสือรับรองข้อมูล) left blank")
    if args.materiality_base is None:
        warnings.append("no --materiality-base given — row 49 (จำนวนที่ใช้เป็นฐาน) left blank; "
                         "requires งบการเงิน figures (later-phase skill)")

    audit_fee_num = float(re.sub(r"[^\d.]", "", audit_fee))
    expense = args.audit_expense if args.audit_expense is not None else round(audit_fee_num * 0.3 / 100) * 100
    if args.audit_expense is None:
        warnings.append(f"ประมาณการค่าใช้จ่าย auto-estimated at {expense:g} (30% of audit_fee, rounded) — "
                         "review and override with --audit-expense if wrong")

    wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
    ws = wb["ข้อมูลลูกค้า"]

    def setb(row, value):
        ws.cell(row=row, column=2).value = value

    setb(1, fiscal_year)
    setb(2, company)
    setb(3, entity_label)
    addr = clean(prof.get("address"))
    if addr:
        setb(4, expand_address(addr))
    setb(5, tax_id)
    business_desc = clean(args.business_desc) if args.business_desc else business_type_text
    if business_desc:
        # ground truth only prefixes บริษัทฯ ประกอบกิจการ for บริษัท/บริษัทจำกัด clients; หจก. clients state it bare
        setb(6, f"บริษัทฯ ประกอบกิจการ\n{business_desc}" if is_company else business_desc)
    if not args.business_desc and business_type_text:
        warnings.append("ประเภทธุรกิจ (row 6) used the raw DBD business_type text verbatim — "
                         "the firm normally tightens this to one short phrase; pass --business-desc to override")
    if args.approval_date:
        setb(7, be_date(args.approval_date))
    if incorp_date:
        setb(8, incorp_date)
    setb(9, period_start_date)
    setb(10, be_date(period_end))
    if is_liq:
        warnings.append("liquidation client — prior-period rows (11/12) left blank, matching the "
                         "ground-truth liquidation case which shows no prior-year comparison")
    elif is_first_year:
        pass  # already warned above when period_start_date was set to incorporation_date
    else:
        if prior_start:
            setb(11, be_date(prior_start))
        if prior_period_end:
            setb(12, be_date(prior_period_end))
    for i, name in enumerate(directors):
        setb(13 + i, name)
    if args.director_authority:
        setb(16, args.director_authority)
    setb(18, auditor_name)
    license_phrase = (f"ผู้สอบบัญชีรับอนุญาต เลขทะเบียน {license_no}" if auditor_type.startswith("CPA")
                       else f"ผู้สอบบัญชีภาษีอากร เลขทะเบียน {license_no}")
    if auditor_type.startswith("TA"):
        warnings.append("auditor_type is TA — license phrase for row 19 has no ground-truth precedent, verify wording")
    setb(19, license_phrase)
    if team:
        setb(20, len(team))
        for i, name in enumerate(team):
            ws.cell(row=21 + i, column=2).value = name
    if args.audit_start_date:
        setb(26, be_date(args.audit_start_date))
    setb(28, audit_fee_num)
    if args.audit_hours is not None:
        setb(29, args.audit_hours)
    setb(30, expense)
    if args.rep_letter_date:
        setb(39, be_date(args.rep_letter_date))
    if args.materiality_basis:
        setb(46, args.materiality_basis)
    else:
        warnings.append("no --materiality-basis given — row 46 (กำหนดจาก) left blank; ground truth has no "
                         "reliable majority (roughly 1:2 สินทรัพย์รวม:รายได้รวม, ~19% left blank outright) — "
                         "pass --materiality-basis \"สินทรัพย์รวม\" or \"รายได้รวม\" based on audit judgment")
    setb(47, args.materiality_reason or boiler["materiality_reason"])
    setb(48, f"ณ วันที่ {period_end}")
    if args.materiality_base is not None:
        setb(49, args.materiality_base)
    setb(54, args.prior_adjustment_note or boiler["prior_adjustment_note"])
    setb(56, args.biz_environment or boiler["biz_environment"])

    ws301 = wb["301"]
    if boiler["301_row28"]:
        for col, val in enumerate(boiler["301_row28"], start=1):
            ws301.cell(row=28, column=col).value = val

    ws_tb = wb["203 TB (2)"]
    for row, vals in boiler["tb_rows"].items():
        for col, val in enumerate(vals, start=2):
            if val is not None:
                ws_tb.cell(row=row, column=col).value = val
    if args.tb_note:
        ws_tb.cell(row=5, column=2).value = args.tb_note

    out_dir = ctx_path.parent / "WP"
    out_dir.mkdir(parents=True, exist_ok=True)
    y_short = str(y)[-2:]
    fname = f"1 Planning {company} {d:02d}{mo:02d}{y_short}.xlsx"
    out_path = out_dir / fname
    wb.save(out_path)

    print(json.dumps({
        "ok": True,
        "output": str(out_path),
        "business_type": business_type,
        "warnings": warnings,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
