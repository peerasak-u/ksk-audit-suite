# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl"]
# ///
"""Dev-only: regenerate the client-neutral Planning.xlsx template from ground truth.

Reads ONE real file under `5_ตัวอย่างไฟล์ผลลัพธ์/` (read-only — never writes there),
clears every client-specific cell in sheet `ข้อมูลลูกค้า` plus the free-text risk/TB
rows in sheets `301` and `203 TB (2)`, and writes the result to
`.claude/skills/audit-planning/assets/planning_template.xlsx`.

All formulas in the other 9 sheets (which cascade off `ข้อมูลลูกค้า`) are left
completely untouched, as are the fixed instructional hints in column C and the
boilerplate rows that never vary (301 rows 1-27, 203 TB(2) header).

Not part of a normal run — only re-run this if the firm's own Planning.xlsx
master template changes shape.

Usage (from repository root):
    uv run .claude/skills/audit-planning/scripts/build_template.py

NOTE (public repo): the source path below points at a real client's folder name
under the firm's private `5_ตัวอย่างไฟล์ผลลัพธ์/` (gitignored, never committed).
Substitute your own local client folder/file names to actually re-run this.
"""
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent.parent.parent
SOURCE = ROOT / "5_ตัวอย่างไฟล์ผลลัพธ์" / "S [103] <local client folder>" / "WP" / "1 Planning <local client name>.xlsx"
OUT = Path(__file__).resolve().parent.parent / "assets" / "planning_template.xlsx"

# Rows in ข้อมูลลูกค้า to clear (column B only — column C holds the sheet's own
# permanent instructional examples and must never be touched). Rows holding a
# formula (27, 37) are left alone; rows that are firm-wide constants are left
# alone too (they are baked into the template, not filled per client) — see
# references/planning-structure.md for the full row map.
CLEAR_ROWS = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16,
              18, 19, 20, 21, 22, 26, 28, 29, 30, 39, 45, 46, 47, 48, 49, 54, 56]

def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"source ground-truth file not found: {SOURCE}")
    wb = openpyxl.load_workbook(SOURCE, data_only=False)

    ws = wb["ข้อมูลลูกค้า"]
    for row in CLEAR_ROWS:
        ws.cell(row=row, column=2).value = None

    # 301: keep rows 1-27 (fixed header + the two always-identical risk items),
    # clear row 28 (the one variant risk item — script fills per business_type).
    ws301 = wb["301"]
    for col in range(1, 6):
        ws301.cell(row=28, column=col).value = None

    # 203 TB (2): clear the free-text analysis rows 5-10, keep the header rows.
    ws_tb = wb["203 TB (2)"]
    for row in range(5, 11):
        for col in range(1, 5):
            ws_tb.cell(row=row, column=col).value = None

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
