# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl"]
# ///
"""Deterministic DB-join for audit-ingest.

Given an input client folder (under 4_ตัวอย่างไฟล์ลูกค้า/), this:
  1. Parses the job number from the folder name.
  2. Joins the three root Database files (job registry, auditor registry, fee ratecard).
  3. Fills the pinned CONTEXT template with DB-sourced values.
  4. Writes 6_ผลจากสกิล/<folder>/CONTEXT.md and creates the WP/ skeleton dir.
  5. Prints a JSON summary (validation verdict + list of ⟨FILL⟩ markers left) to stdout.

The agent must NOT re-type any DB value; it only resolves the ⟨FILL⟩ markers.

Usage:
    uv run .claude/skills/audit-ingest/scripts/join_databases.py "<path to client folder under 4_...>"

Run from the repository root.
"""
import csv
import json
import re
import sys
import time
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl

DBD_URL = "https://openapi.dbd.go.th/api/v1/juristic_person/{}"
THAI_MONTHS = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
               "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]


def _buddhist_year(y: int) -> int:
    return y if y > 2400 else y + 543


def fmt_thai_long_date(v) -> str:
    """Format a DB cell as Thai long date '16 กุมภาพันธ์ 2569'. Datetimes and strings
    stored inconsistently across rows both normalize here."""
    if isinstance(v, datetime):
        return f"{v.day} {THAI_MONTHS[v.month]} {_buddhist_year(v.year)}"
    return str(v).strip()


def fmt_slash_date(v) -> str:
    """Format a DB cell as DD/MM/BBBB (Buddhist). Passes through existing DD/MM/YYYY strings."""
    if isinstance(v, datetime):
        return f"{v.day:02d}/{v.month:02d}/{_buddhist_year(v.year)}"
    return str(v).strip()

JOB_DB = "Database งาน.xlsx"
AUDITOR_DB = "Database ข้อมูลผู้สอบ.csv.xlsx"
RATE_DB = "Database เรทคิดเงิน Audit.csv"
OUT_ROOT = "6_ผลจากสกิล"
INPUT_ROOT = "4_ตัวอย่างไฟล์ลูกค้า"
ARCHIVE_EXT = {".rar", ".zip", ".7z"}


def fail(msg: str) -> None:
    print(json.dumps({"ok": False, "error": msg}, ensure_ascii=False))
    sys.exit(1)


def parse_job_number(folder_name: str) -> str | None:
    m = re.search(r"\[(\d+)\]", folder_name)
    if m:
        return m.group(1)
    m = re.search(r"งาน\s*(\d+)", folder_name)
    return m.group(1) if m else None


def load_job_rows(root: Path, job: str) -> list[dict]:
    """Return ALL rows whose job number matches (job numbers are NOT unique in this DB)."""
    wb = openpyxl.load_workbook(root / JOB_DB, read_only=True, data_only=True)
    ws = wb["รวมรายชื่องบการเงิน"]
    header, matches = None, []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c is not None else "" for c in row]
            continue
        if row[0] is not None and str(row[0]).strip() == job:
            matches.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
    return matches


def select_job_row(rows: list[dict], folder_name: str) -> tuple[dict | None, str]:
    """Disambiguate duplicate job rows by matching company + referrer against the folder.

    Returns (chosen_row_or_None, note). None means ambiguous — caller must flag REVIEW.
    """
    if len(rows) == 1:
        return rows[0], ""
    folder_tok = distinctive_tokens(folder_name)
    scored = []
    for r in rows:
        cand_tok = distinctive_tokens(str(r.get("ชื่อกิจการ") or "")) | distinctive_tokens(str(r.get("ที่มา") or ""))
        scored.append((len(folder_tok & cand_tok), r))
    scored.sort(key=lambda x: x[0], reverse=True)
    top = scored[0][0]
    winners = [r for s, r in scored if s == top]
    labels = " vs ".join(f"'{r.get('ชื่อกิจการ')}' ({r.get('ที่มา')})" for _, r in scored)
    if top == 0 or len(winners) > 1:
        return None, f"{len(rows)} rows share this job number and none/both match the folder: {labels}"
    return winners[0], f"disambiguated from {len(rows)} rows sharing this job number ({labels})"


def load_auditor(root: Path, first_name: str) -> dict | None:
    if not first_name:
        return None
    wb = openpyxl.load_workbook(root / AUDITOR_DB, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    idx = {str(h).strip(): i for i, h in enumerate(header) if h}
    ni = idx.get("ชื่อ")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if ni is not None and row[ni] and str(row[ni]).strip() == first_name.strip():
            return {str(h).strip(): row[i] for i, h in enumerate(header) if h}
    return None


def load_ratecard(root: Path, referrer: str, revenue) -> str:
    """Return expected 'ระบุในงบ' fee for the revenue bracket, or '' if not resolvable."""
    try:
        rev = float(revenue)
    except (TypeError, ValueError):
        return ""
    rev_m = rev / 1_000_000
    rows = []
    with open(root / RATE_DB, encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            rows.append(r)
    # Prefer rows matching the referrer; fall back to all rows.
    cand = [r for r in rows if referrer and referrer in (r.get("ที่มา") or "")] or rows
    best = None
    for r in cand:
        name = (r.get("Name") or "").strip()
        try:
            thresh = float(r.get("รายได้") or 0)
        except ValueError:
            thresh = 0
        fee = r.get("ระบุในงบ") or ""
        if name == "งบเปล่า":
            if rev_m == 0:
                return fee
            continue
        # "รายได้ไม่เกิน <thresh> ล้านบาท"
        if rev_m <= thresh:
            if best is None or thresh < best[0]:
                best = (thresh, fee)
    return best[1] if best else ""


def dbd_lookup(tax_id: str, attempts: int = 3) -> dict | None:
    """Query the DBD OpenAPI (public, no auth) by 13-digit tax id. Retries with backoff
    because the endpoint throttles rapid calls. None on persistent failure."""
    if not re.fullmatch(r"\d{13}", tax_id or ""):
        return None
    for i in range(attempts):
        try:
            with urllib.request.urlopen(DBD_URL.format(tax_id), timeout=25) as r:
                data = json.load(r)
            return data["data"][0]["cd:OrganizationJuristicPerson"]
        except (KeyError, IndexError, TypeError):
            return None  # reached DBD but no such juristic person
        except Exception:
            if i < attempts - 1:
                time.sleep(1.5 * (i + 1))
    return None


def dbd_fields(o: dict) -> dict:
    """Map a DBD juristic-person record to CONTEXT field values."""
    def g(k: str) -> str:
        v = o.get(k)
        return str(v).strip() if v not in (None, "") else ""

    reg = g("cd:OrganizationJuristicRegisterDate")  # YYYYMMDD, Gregorian
    inc = f"{reg[6:8]}/{reg[4:6]}/{reg[0:4]}" if len(reg) == 8 and reg.isdigit() else "—"
    typ = g("cd:OrganizationJuristicType")
    entity = ("หจก. (ห้างหุ้นส่วนจำกัด)" if "ห้างหุ้นส่วน" in typ
              else "บจ. (บริษัทจำกัด)" if "บริษัท" in typ else "")
    bt = ""
    try:
        bt = o["cd:OrganizationJuristicObjective"]["td:JuristicObjective"]["td:JuristicObjectiveTextTH"] or ""
    except (KeyError, TypeError):
        bt = ""
    return {
        "name": g("cd:OrganizationJuristicNameTH"),
        "en": g("cd:OrganizationJuristicNameEN"),
        "status": g("cd:OrganizationJuristicStatus"),
        "capital": fmt_num(g("cd:OrganizationJuristicRegisterCapital")),
        "incorporation_date": inc,
        "business_type": bt.strip(),
        "address": compose_dbd_address(o.get("cd:OrganizationJuristicAddress")),
        "entity_type": entity,
    }


def compose_dbd_address(addr: dict | None) -> str:
    try:
        a = addr["cr:AddressType"]
    except (KeyError, TypeError):
        return ""
    def g(k):
        v = a.get(k)
        return str(v).strip() if v not in (None, "") else ""
    def sub(k, key="cr:CitySubDivisionTextTH"):
        v = a.get(k) or {}
        return str(v.get(key, "") or "").strip() if isinstance(v, dict) else ""
    parts = [g("cd:Address")]
    if g("cd:Road"):
        parts.append("ถ." + g("cd:Road"))
    tambon = sub("cd:CitySubDivision")
    amphoe = a.get("cd:City", {}).get("cr:CityTextTH", "") if isinstance(a.get("cd:City"), dict) else ""
    province = a.get("cd:CountrySubDivision", {}).get("cr:CountrySubDivisionTextTH", "") if isinstance(a.get("cd:CountrySubDivision"), dict) else ""
    if tambon:
        parts.append("ต." + tambon)
    if amphoe:
        parts.append("อ." + str(amphoe).replace("เมือง", "เมือง"))
    if province:
        parts.append("จ." + str(province))
    return " ".join(p for p in parts if p).strip()


def derive_entity_type(name: str) -> str:
    n = name or ""
    if "ห้างหุ้นส่วน" in n or "หจก" in n:
        return "หจก. (ห้างหุ้นส่วนจำกัด)"
    return "บจ. (บริษัทจำกัด)"


def prior_period(period_end: str) -> str:
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", period_end or "")
    if not m:
        return "—"
    d, mo, y = m.groups()
    return f"{d}/{mo}/{int(y) - 1}"


def fmt_num(v) -> str:
    if v is None or v == "":
        return "—"
    try:
        f = float(v)
        return f"{int(f):,}" if f == int(f) else f"{f:,.2f}"
    except (TypeError, ValueError):
        return str(v)


def distinctive_tokens(name: str) -> set[str]:
    stop = {"บริษัท", "จำกัด", "บจ", "บจก", "หจก", "ห้างหุ้นส่วน", "จก", ""}
    cleaned = re.sub(r"[.\[\]()0-9]", " ", name or "")
    return {t for t in cleaned.split() if t not in stop and len(t) >= 2}


def build_inventory(client_dir: Path) -> tuple[str, str]:
    files, archives = [], []
    for p in sorted(client_dir.rglob("*")):
        if p.is_file():
            (archives if p.suffix.lower() in ARCHIVE_EXT else files).append(p.name)
    lines = ["| document | readable |", "|---|---|"]
    for f in files:
        note = "scan → use vision" if f.lower().endswith(".pdf") else "—"
        lines.append(f"| {f} | {note} |")
    for a in archives:
        lines.append(f"| {a} (archive) | ⚠ extract with `tar -xf`, dedup, add new files |")
    has_cert = any("รับรอง" in f for f in files)
    missing = "" if has_cert else "- ❌ **missing: หนังสือรับรอง** → director / authorized_signatory stay ⚠ (DBD covers legal name / type / capital / incorporation / business_type by tax_id; the director list is not in DBD)."
    return "\n".join(lines), missing


def main() -> None:
    if len(sys.argv) != 2:
        fail("usage: join_databases.py '<path to client folder under 4_...>'")
    client_dir = Path(sys.argv[1])
    if not client_dir.is_dir():
        fail(f"not a directory: {client_dir}")
    root = Path.cwd()
    for db in (JOB_DB, AUDITOR_DB, RATE_DB):
        if not (root / db).exists():
            fail(f"missing database at repo root: {db} (run from repository root)")

    folder_name = client_dir.name
    job = parse_job_number(folder_name)
    if not job:
        fail(f"cannot parse job number from folder name: {folder_name}")

    candidates = load_job_rows(root, job)
    job_row, dis_note = (None, "")
    if candidates:
        job_row, dis_note = select_job_row(candidates, folder_name)
    v: dict[str, str] = {"job_number": job, "folder_name": folder_name}

    if not candidates or job_row is None:
        # Job absent ([364]) OR ambiguous duplicate that can't be resolved — all-⚠ skeleton.
        FILL = "⚠ ⟨FILL: unresolved in DB — extract from documents⟩"
        v.update({k: FILL for k in (
            "company_name", "referrer", "auditor_first_name", "period_end",
            "entity_type", "tax_id", "revenue", "auditor_name", "auditor_license",
            "auditor_type", "opinion_type", "sign_date", "audit_fee", "ratecard_fee",
            "company_legal_name", "company_name_en", "juristic_status",
            "registered_capital", "incorporation_date", "business_type", "dbd_address",
        )})
        v["prior_period_end"] = "—"
        v["tax_id_status"] = v["auditor_status"] = v["fee_status"] = "⚠"
        v["entity_source"] = v["dbd_source"] = "—"
        v["dbd_status"] = "⚠"
        if not candidates:
            v["validation_verdict"] = "JOB NOT IN DB — extract all metadata from documents, do NOT auto-generate"
            v["validation_notes"] = "- ⚠ Job number not found in Database งาน. Treat every DB field above as unverified."
        else:
            v["validation_verdict"] = "AMBIGUOUS — multiple DB rows share this job number; confirm which before generating"
            v["validation_notes"] = f"- ⚠ {dis_note}"
    else:
        name = str(job_row.get("ชื่อกิจการ") or "").strip()
        referrer = str(job_row.get("ที่มา") or "").strip()
        first = str(job_row.get("ชื่อผู้สอบ") or "").strip()
        revenue = job_row.get("ขนาดกิจการ(รายได้รวม)")
        tax_id = str(job_row.get("เลขนิติ") or "").strip()
        aud = load_auditor(root, first)

        v["company_name"] = name or "—"
        v["referrer"] = referrer or "—"
        v["auditor_first_name"] = first or "—"
        v["tax_id"] = tax_id or "—"

        # DBD OpenAPI lookup by tax id — authoritative full legal name + registry facts.
        dbd = dbd_lookup(tax_id)
        if dbd:
            f = dbd_fields(dbd)
            unreach = "⚠ ⟨FILL: not returned by DBD — from บอจ.5/หนังสือรับรอง⟩"
            v["company_legal_name"] = f["name"] or name
            v["company_name_en"] = f["en"] or "—"
            v["juristic_status"] = f["status"] or "—"
            v["registered_capital"] = f["capital"] if f["capital"] != "—" else unreach
            v["incorporation_date"] = f["incorporation_date"]
            v["business_type"] = f["business_type"] or "—"
            v["dbd_address"] = f["address"] or unreach
            v["entity_type"] = f["entity_type"] or derive_entity_type(name)
            v["entity_source"] = "DBD" if f["entity_type"] else "derived"
            v["dbd_source"] = "DBD (tax_id)"
            v["dbd_status"] = "✔"
        else:
            miss = "⚠ ⟨FILL: DBD not found — from บอจ.5/หนังสือรับรอง⟩"
            v["company_legal_name"] = name  # fall back to DB name (may be abbreviated)
            v["company_name_en"] = "—"
            v["juristic_status"] = "—"
            v["registered_capital"] = miss
            v["incorporation_date"] = miss
            v["business_type"] = miss
            v["dbd_address"] = miss
            v["entity_type"] = derive_entity_type(name)
            v["entity_source"] = "derived"
            v["dbd_source"] = "DBD (not found)"
            v["dbd_status"] = "⚠"
        v["tax_id_status"] = "✔" if len(tax_id) == 13 else "⚠ (not 13 digits — verify)"
        v["period_end"] = fmt_slash_date(job_row.get("รอบปีบัญชี")) or "—"
        v["prior_period_end"] = prior_period(v["period_end"])
        v["revenue"] = fmt_num(revenue)
        v["opinion_type"] = str(job_row.get("รูปแบบหน้ารายงาน") or "—").strip()
        v["sign_date"] = fmt_thai_long_date(job_row.get("วันที่หน้ารายงาน")) or "—"

        if aud:
            full = f"{aud.get('คำนำหน้า') or ''}{aud.get('ชื่อ') or ''} {aud.get('นามสกุล') or ''}".strip()
            cpata = str(aud.get("CPA/ TA") or aud.get("CPA/TA") or "").upper()
            v["auditor_name"] = full or "—"
            v["auditor_license"] = str(aud.get("เลข CPA TA") or "—").strip()
            v["auditor_type"] = "CPA (ผู้สอบบัญชีรับอนุญาต)" if "CPA" in cpata else "TA (ผู้สอบภาษีอากร)"
            v["auditor_status"] = "✔"
        else:
            v["auditor_name"] = v["auditor_license"] = v["auditor_type"] = f"⚠ ⟨FILL: auditor '{first}' not in DB ผู้สอบ⟩"
            v["auditor_status"] = "⚠"

        db_fee = fmt_num(job_row.get("ค่าสอบจากงบการเงิน"))
        rate_fee = load_ratecard(root, referrer, revenue)
        v["audit_fee"] = db_fee
        v["ratecard_fee"] = fmt_num(rate_fee) if rate_fee else "n/a"
        v["fee_status"] = "✔" if (not rate_fee or db_fee == fmt_num(rate_fee)) else "⚠ (DB ≠ ratecard — verify)"

        db_tok, fold_tok = distinctive_tokens(name), distinctive_tokens(folder_name)
        shared = db_tok & fold_tok
        if shared:
            v["validation_verdict"] = "PASS — company name matches folder"
            v["validation_notes"] = f"- ✔ shared token(s): {', '.join(sorted(shared))}"
        else:
            v["validation_verdict"] = "REVIEW — DB company name does NOT match folder name; do NOT auto-generate until confirmed"
            v["validation_notes"] = (f"- ⚠ No shared distinctive token between DB name ('{name}') "
                                     f"and folder ('{folder_name}'). Possible wrong job number.")
        if dis_note:
            v["validation_notes"] += f"\n- ℹ {dis_note}"

    inventory, missing = build_inventory(client_dir)
    v["document_inventory"] = inventory
    v["missing_docs"] = missing

    tpl_path = Path(__file__).resolve().parent.parent / "references" / "context-template.md"
    tpl = tpl_path.read_text(encoding="utf-8")
    for key, val in v.items():
        tpl = tpl.replace("{{" + key + "}}", str(val))
    leftover = re.findall(r"\{\{(\w+)\}\}", tpl)
    if leftover:
        fail(f"template placeholders not filled: {sorted(set(leftover))}")

    out_dir = root / OUT_ROOT / folder_name
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "WP").mkdir(exist_ok=True)
    context_path = out_dir / "CONTEXT.md"
    context_path.write_text(tpl, encoding="utf-8")

    fill_markers = re.findall(r"⟨FILL:[^⟩]*⟩", tpl)
    print(json.dumps({
        "ok": True,
        "job_number": job,
        "found_in_db": job_row is not None,
        "context_path": str(context_path),
        "wp_dir": str(out_dir / "WP"),
        "validation_verdict": v["validation_verdict"],
        "fill_markers_remaining": len(fill_markers),
        "company_name": v.get("company_name"),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
