# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl", "pypdf", "python-docx"]
# ///
"""Pipeline status + format-QA dashboard for a KSK audit client (skill 6, audit-orchestrate).

The five per-client deliverables are produced by five independent skills that chain through a
shared `CONTEXT.md` (see docs/OVERVIEW.md). This skill does NOT re-do any of their work — it
inspects a client's `6_ผลจากสกิล/<client>/` output and reports:

  - which pipeline stage the client is at, and which of the 5 target files exist;
  - the next action (which skill to run) and any human gate that is blocking;
  - a FORMAT-QA check of each file that exists, against the locked format contract for its
    type (not a content diff — content differs per client by design; format must not);
  - a cross-check that the matching ground-truth client folder under `5_ตัวอย่างไฟล์ผลลัพธ์/`
    has the same set of deliverables (so nothing is silently missing).

It is a read-only dashboard: it opens files but never writes them. Drive the actual work with
the individual skills per the runbook in SKILL.md.

Usage (from repository root):
    uv run .claude/skills/audit-orchestrate/scripts/orchestrate.py "PATH/TO/CONTEXT.md"
    uv run .claude/skills/audit-orchestrate/scripts/orchestrate.py "PATH/TO/6_.../<client>"
    uv run .claude/skills/audit-orchestrate/scripts/orchestrate.py --all

Prints JSON: one report object, or {"clients": [...]} for --all.
"""
import argparse
import json
import re
import sys
from pathlib import Path

OUT_BASE = "6_ผลจากสกิล"
GT_BASE = "5_ตัวอย่างไฟล์ผลลัพธ์"

# The 5 target deliverables, the skill that makes each, and how to recognise the file by name.
# order = pipeline order for display; phase = A (pre-judgment) / B (post human WP adjustment).
TARGETS = [
    {"key": "planning", "label": "1 Planning.xlsx", "skill": "audit-planning", "phase": "A"},
    {"key": "report", "label": "2 หน้ารายงาน.docx", "skill": "audit-cover-report", "phase": "A"},
    {"key": "cover", "label": "ใบปะหน้างบการเงิน.docx", "skill": "audit-cover-report", "phase": "A"},
    {"key": "workpaper", "label": "4 งบการเงิน.xlsx", "skill": "audit-workpaper", "phase": "A"},
    {"key": "cit50", "label": "3 CIT50.pdf", "skill": "audit-cit50", "phase": "B"},
    {"key": "financials", "label": "4 งบการเงิน (final).xlsx", "skill": "audit-financials", "phase": "B"},
]

# locked format invariants
PLANNING_SHEETS = ["ข้อมูลลูกค้า", "ประเภทบัญชี", "000", "001", "101", "102", "103",
                   "202", "203 TB (2)", "301", "302", "401", "601", "608"]
FS_CORE_SHEETS = ["งบการเงิน", "TB", "Mapping", "ปรับปรุง", "ภาษีเงินได้"]


def classify(name: str) -> str | None:
    """Map a filename to a target key by keyword (priority order matters)."""
    n = name
    low = n.lower()
    if n.startswith("~$"):
        return None
    if "ใบปะหน้า" in n:
        return "cover"
    if "หน้ารายงาน" in n or n.strip().startswith("2 "):
        return "report"
    if "planning" in low or n.strip().startswith("1 "):
        return "planning"
    if "cit50" in low or "ภ.ง.ด" in n or "ภงด" in n or n.strip().startswith("3 "):
        return "cit50"
    if n.lower().endswith((".xlsx", ".xls")) and ("งบการเงิน" in n or "ร่างงบ" in n or n.strip().startswith("4 ")):
        return "final" if "(final)" in low else "workpaper"
    return None


def find_wp_dir(client_dir: Path) -> Path:
    """The WP/ subfolder holding deliverables (5_ uses 'WP <year>'; 6_ uses 'WP')."""
    subs = [d for d in client_dir.iterdir() if d.is_dir() and d.name.startswith("WP")]
    for d in subs:  # prefer one that actually contains files
        if any(f.is_file() for f in d.iterdir()):
            return d
    return subs[0] if subs else client_dir


def scan_deliverables(wp_dir: Path) -> dict[str, list[Path]]:
    found: dict[str, list[Path]] = {}
    if not wp_dir.exists():
        return found
    for f in sorted(wp_dir.rglob("*")):
        if not f.is_file():
            continue
        k = classify(f.name)
        if k == "final":
            k = "financials"
        if k:
            found.setdefault(k, []).append(f)
    return found


# ── format-QA per file type (checks the LOCKED format, not content) ───────────────────
def qa_xlsx_sheets(path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    return wb.sheetnames


def qa_planning(path: Path) -> dict:
    try:
        sheets = qa_xlsx_sheets(path)
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "notes": [f"cannot open: {e}"]}
    missing = [s for s in PLANNING_SHEETS if s not in sheets]
    return {"ok": not missing,
            "notes": ([f"missing fixed sheets: {', '.join(missing)}"] if missing else ["14 fixed sheets present"])}


def qa_workpaper(path: Path, *, expect_final=False) -> dict:
    import openpyxl
    try:
        sheets = openpyxl.load_workbook(path, read_only=True).sheetnames
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "notes": [f"cannot open: {e}"]}
    notes = []
    missing = [s for s in FS_CORE_SHEETS if s not in sheets]
    ok = not missing
    if missing:
        notes.append(f"missing core FS sheets: {', '.join(missing)}")
    else:
        notes.append("core FS sheets present")
    # human-gate signal: has the WP been opened & saved in Excel (cached values present)?
    dn = wb.defined_names.get("FS_TOTAL_ASSETS_CY")
    adjusted = None
    if dn is not None:
        for sh, coord in dn.destinations:
            adjusted = wb[sh][coord.replace("$", "")].value is not None
            break
    if adjusted is False:
        notes.append("⚠ ยังไม่ถูกเปิด/บันทึกใน Excel (ยังไม่ได้ปรับปรุง/ยังไม่มีค่าคำนวณ)")
    elif adjusted is True:
        notes.append("มีค่าคำนวณแล้ว (ถูกเปิด/บันทึกใน Excel)")
    return {"ok": ok, "notes": notes, "adjusted": adjusted}


def qa_docx(path: Path) -> dict:
    try:
        import docx
        d = docx.Document(str(path))
        paras = [p for p in d.paragraphs if p.text.strip()]
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "notes": [f"cannot open: {e}"]}
    return {"ok": len(paras) > 0,
            "notes": [f"{len(paras)} non-empty paragraphs" if paras else "document is empty"]}


def qa_cit50(path: Path) -> dict:
    try:
        from pypdf import PdfReader
        r = PdfReader(str(path))
        fields = r.get_fields() or {}
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "notes": [f"cannot open/parse AcroForm: {e}"]}
    n = len(fields)
    return {"ok": n > 900, "notes": [f"AcroForm with {n} fields" + ("" if n > 900 else " (expected ~925)")]}


QA = {"planning": qa_planning, "report": qa_docx, "cover": qa_docx,
      "cit50": qa_cit50, "workpaper": qa_workpaper,
      "financials": lambda p: qa_workpaper(p, expect_final=True)}


# ── ground-truth matching ─────────────────────────────────────────────────────────────
def job_no(name: str):
    m = re.search(r"\[(\d+)\]", name)
    return m.group(1) if m else None


def match_ground_truth(client_dir: Path, repo: Path) -> Path | None:
    gt_base = repo / GT_BASE
    if not gt_base.exists():
        return None
    name = client_dir.name
    exact = gt_base / name
    if exact.exists():
        return exact
    jn = job_no(name)
    if jn:
        for d in gt_base.iterdir():
            if d.is_dir() and job_no(d.name) == jn:
                return d
    return None


# ── per-client report ─────────────────────────────────────────────────────────────────
def report_client(client_dir: Path, repo: Path) -> dict:
    ctx = client_dir / "CONTEXT.md"
    wp = find_wp_dir(client_dir)
    found = scan_deliverables(wp)

    files = {}
    adjusted = None
    for t in TARGETS:
        k = t["key"]
        paths = found.get(k, [])
        entry = {"label": t["label"], "skill": t["skill"], "phase": t["phase"],
                 "present": bool(paths)}
        if paths:
            entry["path"] = str(paths[-1].relative_to(repo))
            qa = QA[k](paths[-1])
            entry["format_ok"] = qa["ok"]
            entry["format_notes"] = qa["notes"]
            if k == "workpaper":
                adjusted = qa.get("adjusted")
        files[k] = entry

    # ground-truth deliverable set
    gt = match_ground_truth(client_dir, repo)
    gt_info = None
    if gt:
        gt_found = scan_deliverables(find_wp_dir(gt))
        gt_info = {"folder": gt.name, "has": sorted(gt_found.keys())}

    # pipeline stage + next actions
    actions: list[str] = []
    if not ctx.exists():
        stage = "0-not-ingested"
        actions.append("run audit-ingest on the client folder to produce CONTEXT.md + WP/ skeleton")
    else:
        phase_a_keys = ["planning", "report", "cover", "workpaper"]
        missing_a = [k for k in phase_a_keys if not files[k]["present"]]
        if missing_a:
            stage = "1-phase-A-scaffolding"
            for k in missing_a:
                actions.append(f"run {files[k]['skill']} → {files[k]['label']}")
        elif adjusted is not True:
            stage = "2-awaiting-human-WP-adjust"
            actions.append("HUMAN GATE: open '4 งบการเงิน' in Excel, adjust TB, classify every "
                           "Mapping!H, post AJEs, save — then continue phase B")
        else:
            missing_b = [k for k in ("financials", "cit50") if not files[k]["present"]]
            if missing_b:
                stage = "3-phase-B-finalizing"
                for k in missing_b:
                    actions.append(f"run {files[k]['skill']} → {files[k]['label']}")
            else:
                stage = "4-complete"
                actions.append("all 5 deliverables produced — do final human review vs ground truth")

    # roll-up flags
    fmt_fail = [files[k]["label"] for k in files
                if files[k]["present"] and not files[k].get("format_ok", True)]

    return {
        "client": client_dir.name,
        "context_present": ctx.exists(),
        "stage": stage,
        "wp_dir": str(wp.relative_to(repo)) if wp.exists() else None,
        "files": files,
        "format_failures": fmt_fail,
        "ground_truth": gt_info,
        "next_actions": actions,
    }


def resolve_clients(arg: str | None, all_flag: bool, repo: Path) -> list[Path]:
    if all_flag:
        base = repo / OUT_BASE
        if not base.exists():
            return []
        return sorted(d for d in base.iterdir()
                      if d.is_dir() and ((d / "CONTEXT.md").exists() or (d / "WP").exists()))
    p = Path(arg)
    if not p.is_absolute():
        p = repo / p
    if p.is_file() and p.name == "CONTEXT.md":
        return [p.parent]
    if p.is_dir():
        return [p]
    print(json.dumps({"ok": False, "error": f"not a CONTEXT.md or client folder: {arg}"}, ensure_ascii=False))
    sys.exit(1)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("target", nargs="?", help="path to a client's CONTEXT.md or client folder")
    ap.add_argument("--all", action="store_true", help="report every client under 6_ผลจากสกิล/")
    ap.add_argument("--repo", default=".", help="repository root (default: cwd)")
    args = ap.parse_args()
    if not args.target and not args.all:
        ap.error("give a CONTEXT.md / client folder, or --all")

    repo = Path(args.repo).resolve()
    clients = resolve_clients(args.target, args.all, repo)
    if not clients:
        print(json.dumps({"ok": True, "clients": [], "note": "no clients found under 6_ผลจากสกิล/"},
                         ensure_ascii=False))
        return

    reports = [report_client(c, repo) for c in clients]
    if args.all or len(reports) > 1:
        print(json.dumps({"ok": True, "clients": reports}, ensure_ascii=False))
    else:
        print(json.dumps({"ok": True, **reports[0]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
