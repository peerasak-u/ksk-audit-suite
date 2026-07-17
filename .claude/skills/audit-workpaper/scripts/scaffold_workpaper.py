# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl"]
# ///
"""Scaffold the audit working-paper workbook (`4 งบการเงิน <client>.xlsx`) for a client.

This is skill 5 (audit-workpaper). It writes the *structure* of the WP before the numbers
and judgment exist — a formula-linked งบการเงิน template, a Mapping (account → FS caption)
sheet, an adjusted-TB layout, an empty adjusting-entries grid, and a tax shell — so that
once the auditor classifies each account in `Mapping!H` and adjusts the TB, the financial
statements recompute themselves ("approach A"). It NEVER invents numbers or classifications.

The shared structure this writes (and skill 5.5 `audit-financials` later reads) is specified
in `docs/financials-contract.md`. v1 covers บจ. going-concern only.

Usage (from repository root):
    uv run .claude/skills/audit-workpaper/scripts/scaffold_workpaper.py \
        "PATH/TO/6_ผลจากสกิล/<client>/CONTEXT.md" [--client-tb PATH [--tb-sheet NAME]]

Prints JSON: output path + warnings (every judgment cell left blank, every assumption made).
"""
import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

FS = "งบการเงิน"          # financial-statements sheet name
MAP = "Mapping"
TB = "TB"
AJE = "ปรับปรุง"
TAX = "ภาษีเงินได้"
CAPREF = "_captions"      # hidden helper sheet holding the controlled vocabulary
NUMFMT = r"_(* #,##0.00_);_(* \(#,##0.00\);_(* \-??_);_(@_)"   # accounting format (contract §10.1)

# ── Print contract (docs/financials-contract.md §10) ──────────────────────────────
# The งบ is printed and bound behind the cover + auditor's report, so it is page 4 onward.
# These values are a LOCKED format taken from all 10 ground-truth files, not preferences.
DEFAULT_FONT = "Cordia New"      # workbook Normal style
FS_FONT = "Browallia New"        # every cell on the งบการเงิน sheet
FS_SIZE = 14
FOOTER_FONT = "Angsana New"
ROW_H = 21.0                     # uniform row height — what makes the page grid predictable
ROWS_PER_PAGE = 36               # 36 × 21pt = 756pt ≤ 771pt printable height of A4
FIRST_PAGE_NUMBER = 4            # ใบปะหน้า + หน้ารายงานผู้สอบ occupy pages 1–3
COL_W = {"A": 31.7, "B": 16.7, "C": 16.7, "D": 0.9, "E": 16.7, "F": 0.9, "G": 16.7}
CHECK_COL = 9                    # column I — deliberately outside the $A:$G print area

# ── Controlled caption vocabulary (docs/financials-contract.md §4) ────────────────
# side = which Mapping column carries the natural balance: "C" = ยอด เดบิต, "D" = ยอด เครดิต.
CUR_ASSETS = [
    ("เงินสดและรายการเทียบเท่าเงินสด", "C"),
    ("ลูกหนี้การค้าและลูกหนี้หมุนเวียนอื่น", "C"),
    ("เงินให้กู้ยืมระยะสั้น", "C"),
    ("สินค้าคงเหลือ", "C"),
    ("สินทรัพย์หมุนเวียนอื่น", "C"),
]
NONCUR_ASSETS = [
    ("ที่ดิน อาคารและอุปกรณ์", "C"),
    ("สินทรัพย์ไม่มีตัวตน", "C"),
    ("สินทรัพย์ไม่หมุนเวียนอื่น", "C"),
]
CUR_LIAB = [
    ("เจ้าหนี้การค้าและเจ้าหนี้หมุนเวียนอื่น", "D"),
    ("เงินกู้ยืมระยะสั้น", "D"),
    ("ภาษีเงินได้ค้างจ่าย", "D"),
]
NONCUR_LIAB = [
    ("เงินกู้ยืมระยะยาว", "D"),
    ("หนี้สินไม่หมุนเวียนอื่น", "D"),
]
SHARE_CAP = ("ทุนที่ออกและชำระแล้ว", "D")
RETAINED = ("กำไร(ขาดทุน)สะสม", "D")
REVENUE = [
    ("รายได้จากการขายหรือบริการ", "D"),
    ("รายได้อื่น", "D"),
]
EXPENSES = [
    ("ต้นทุนขายหรือบริการ", "C"),
    ("ค่าใช้จ่ายในการขายและบริหาร", "C"),
    ("ต้นทุนทางการเงิน", "C"),
]

# All captions that get a live SUMIF note-detail row (order defines note numbering).
ALL_CAPTIONS = (
    [c for c, _ in CUR_ASSETS + NONCUR_ASSETS]
    + [c for c, _ in CUR_LIAB + NONCUR_LIAB]
    + [SHARE_CAP[0], RETAINED[0]]
    + [c for c, _ in REVENUE + EXPENSES]
)
SIDE = dict(CUR_ASSETS + NONCUR_ASSETS + CUR_LIAB + NONCUR_LIAB + [SHARE_CAP, RETAINED] + REVENUE + EXPENSES)

# Fixed accounting-policy boilerplate (docs/financials-contract.md §7).
POLICY_BASIS = (
    "งบการเงินนี้จัดขึ้นตามมาตรฐานการรายงานทางการเงินสำหรับกิจการที่ไม่มีส่วนได้เสียสาธารณะ "
    "ที่ออกโดยสภาวิชาชีพบัญชี และวิธีปฏิบัติทางการบัญชีที่รับรองทั่วไปในประเทศไทย "
    "แสดงรายการตามประกาศของกรมพัฒนาธุรกิจการค้า ลงวันที่ 27 ตุลาคม พ.ศ.2566 "
    "เรื่องกำหนดรายการย่อที่ต้องมีในงบการเงิน พ.ศ.2566 "
    "งบการเงินนี้จัดทำบัญชีตามกฎหมายเป็นภาษาไทยและเป็นเงินบาท ใช้เกณฑ์ราคาทุนเดิมในการวัดมูลค่า"
)
POLICY_ITEMS = [
    ("เงินสดและรายการเทียบเท่าเงินสด",
     "เงินสดและรายการเทียบเท่าเงินสด ประกอบด้วย เงินสดในมือและเงินฝากธนาคารกระแสรายวันและออมทรัพย์ "
     "ที่ครบกำหนดไม่เกิน 3 เดือน ไม่รวมรายการเงินฝากธนาคารที่ภาระค้ำประกัน"),
    ("สินค้าคงเหลือ",
     "สินค้าคงเหลือแสดงด้วยราคาทุนหรือมูลค่าสุทธิที่จะได้รับแล้วแต่ราคาใดจะต่ำกว่า"),
    ("ที่ดิน อาคารและอุปกรณ์",
     "ที่ดินบันทึกด้วยราคาทุนเริ่มแรก ที่ดินไม่มีการคิดค่าเสื่อมราคาเนื่องจากมีอายุการใช้งานไม่จำกัด "
     "อาคารและอุปกรณ์แสดงด้วยราคาทุนหลังหักค่าเสื่อมราคาสะสม ค่าเสื่อมราคาคำนวณโดยวิธีเส้นตรง"),
    ("การรับรู้รายได้และค่าใช้จ่าย",
     "⚠ ⟨FILL: revenue-recognition first sentence differs by business type — trading vs service⟩"),
    ("ภาษีเงินได้นิติบุคคล",
     "การคำนวณภาษีเงินได้นิติบุคคลตามหลักเกณฑ์ประมวลรัษฎากรในอัตราที่กฎหมายกำหนด "
     "ของกำไรสุทธิก่อนค่าใช้จ่ายภาษีเงินได้"),
]


def fail(msg: str) -> None:
    print(json.dumps({"ok": False, "error": msg}, ensure_ascii=False))
    sys.exit(1)


def parse_context(path: Path) -> tuple[dict, dict]:
    """Return (value_by_key, confidence_by_key) from the CONTEXT field tables.

    Rows are `| key | value | source | confidence |`. The confidence column (✔/⚠) is what
    tells us whether a value is safe to PRINT on a client deliverable — the value column
    alone can look clean while the row is still flagged unresolved.
    """
    profile, confidence = {}, {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.lstrip().startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) < 2 or not re.fullmatch(r"[a-z_0-9]+", cells[0]):
            continue
        profile[cells[0]] = cells[1]
        confidence[cells[0]] = cells[-1] if len(cells) >= 4 else ""
    return profile, confidence


_ANNOT_RE = re.compile(r"\s*\([^)]*\)\s*$")


def director_name(profile: dict, confidence: dict, warnings: list) -> str | None:
    """Director name for the sign-off block, or None if it is not safely known.

    This name gets PRINTED on a client deliverable, so an unresolved, low-confidence, or
    role-unconfirmed value yields a blank dotted line instead of a guess (contract §10.4).
    """
    val = clean(profile.get("director_1"))
    if not val:
        warnings.append("CONTEXT director_1 ยังไม่ resolve — บล็อกลงนามเว้น ( .......... ) ไว้ให้กรอกมือ")
        return None
    if "⚠" in (confidence.get("director_1") or ""):
        warnings.append(f"CONTEXT director_1 ({val!r}) confidence = ⚠ — บล็อกลงนามเว้นว่างไว้ ไม่เดาชื่อกรรมการ")
        return None
    # strip the trailing annotation audit-ingest attaches, e.g. "(กรรมการอีกคน: ...)"
    return _ANNOT_RE.sub("", val).strip() or None


def context_verdict(path: Path) -> str | None:
    """Return the §2 validation Verdict line's text, or None if absent."""
    m = re.search(r"Verdict:\s*(.+)", path.read_text(encoding="utf-8"))
    return m.group(1).strip(" *") if m else None


def clean(v):
    if not v or str(v).strip() in {"—", "-"} or "⚠" in str(v) or "⟨FILL" in str(v):
        return None
    return re.sub(r"\s{2,}", " ", str(v).strip())


def period_str(profile: dict):
    """Return (label, filename_tag) from period_end dd/mm/bbbb."""
    pe = profile.get("period_end", "")
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", pe)
    if not m:
        return "⚠ ⟨FILL: period_end⟩", "period"
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    months = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
              "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
    label = f"{d} {months[mo]} พ.ศ. {y}"
    return label, f"{d:02d}{mo:02d}{str(y)[-2:]}"


def be_year(period: str):
    m = re.search(r"/(\d{4})\s*$", period or "")
    return int(m.group(1)) if m else None


# ── client TB import (best-effort) ────────────────────────────────────────────────
# Header text that identifies the real *closing*-balance column pair, as opposed to
# opening (ยกมา/ต้นงวด) or movement (เคลื่อนไหว/ยอด) pairs that some client exports also
# carry (e.g. a ยกมา/เคลื่อนไหว/ยกไป 8-column layout, or a ต้นงวด/ยอด/ปลายงวด 3-pair layout).
_CLOSE_HDR_RE = re.compile(r"(ปลายงวด|ยกไป|คงเหลือ)", re.I)
_DR_HDR_RE = re.compile(r"(เดบิต|เดบิท|debit|\bdr\b)", re.I)
_CR_HDR_RE = re.compile(r"(เครดิต|credit|\bcr\b)", re.I)


def find_closing_pair(rows: list, scan_rows: int = 20):
    """Locate the real closing-balance Dr/Cr column pair by header text.

    Returns (dr_col, cr_col, dr_header_text, cr_header_text) using 0-based column
    indices, or None if no such header pair is found (caller falls back to the
    last-two-numeric-cells heuristic). Handles two real layouts:
      - single-row header where the pair is named together, e.g. "ยกไป Dr" / "ยกไป Cr"
        sitting in adjacent cells;
      - a merged two-row header where a group label ("ยอดคงเหลือ" / "ปลายงวด") sits above
        a เดบิต/เครดิต sub-header row, in the SAME two columns as the group label.
    """
    limit = min(scan_rows, len(rows))
    for i in range(limit):
        row = rows[i]
        if row is None:
            continue
        for j, cell in enumerate(row):
            if not isinstance(cell, str) or not _CLOSE_HDR_RE.search(cell):
                continue
            # case A: this cell itself names Dr, and the very next cell names Cr
            if _DR_HDR_RE.search(cell) and j + 1 < len(row):
                nxt = row[j + 1]
                if isinstance(nxt, str) and _CR_HDR_RE.search(nxt):
                    return j, j + 1, cell.strip(), nxt.strip()
            # case B: this is a merged group-label cell; the เดบิต/เครดิต sub-header sits
            # one or two rows below, in the same two columns (merged-cell layout)
            for k in range(i + 1, min(i + 3, len(rows))):
                subrow = rows[k]
                if subrow is None or j + 1 >= len(subrow):
                    continue
                c_dr, c_cr = subrow[j], subrow[j + 1]
                if (isinstance(c_dr, str) and _DR_HDR_RE.search(c_dr)
                        and isinstance(c_cr, str) and _CR_HDR_RE.search(c_cr)):
                    return j, j + 1, c_dr.strip(), c_cr.strip()
    return None


def import_client_tb(path: Path, sheet_name: str | None, warnings: list) -> list[tuple]:
    """Return list of (account_code, account_name, ending_dr, ending_cr).

    Client GL exports vary wildly (docs/financials-contract.md §2/§8). The closing Dr/Cr
    column pair is located by header text first (ปลายงวด/ยกไป/คงเหลือ, paired with
    เดบิต/เครดิต or Dr/Cr); only when no such header is found does it fall back to
    assuming the last two numeric cells in a row are the closing pair. On any doubt it
    warns and returns what it could parse so the human can correct the TB/Mapping by
    hand — it never fabricates balances.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        warnings.append(f"could not open --client-tb ({e}); produced an empty TB layout to fill by hand")
        return []
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    if sheet_name and sheet_name not in wb.sheetnames:
        warnings.append(f"--tb-sheet '{sheet_name}' not found; used first sheet '{ws.title}'")

    rows = list(ws.iter_rows(values_only=True))
    # find a header row mentioning an account-code-ish column
    header_idx = None
    for i, r in enumerate(rows[:15]):
        joined = " ".join(str(c) for c in r if c is not None)
        if re.search(r"(เลขที่บัญชี|รหัส|ผังบัญชี|บัญชี|account)", joined, re.I):
            header_idx = i
            break
    if header_idx is None:
        warnings.append("could not locate a header row in --client-tb; produced an empty TB to fill by hand")
        return []

    closing = find_closing_pair(rows)
    dr_idx = cr_idx = dr_hdr = cr_hdr = None
    if closing:
        dr_idx, cr_idx, dr_hdr, cr_hdr = closing

    out = []
    seen = 0
    for r in rows[header_idx + 1:]:
        if r is None or all(c is None for c in r):
            continue
        seen += 1
        # account code: extract the LEADING numeric run from cell 0, not a whole-cell
        # match — client cells combine code+suffix ("111101 - CSH001") or code+name
        # ("1001000: เงินสด") in a single cell, and a whole-string match drops those rows.
        raw0 = str(r[0]).strip() if r[0] is not None else ""
        code_m = re.match(r"^(\d[\d\-\.]*)", raw0)
        if not code_m:
            continue
        code = code_m.group(1)
        remainder0 = raw0[code_m.end():].strip(" \t-:.").strip()
        name = ""
        for c in r[1:4]:
            if isinstance(c, str) and c.strip() and not re.match(r"^[\d,\.\-]+$", c.strip()):
                name = c.strip()
                break
        if not name and remainder0 and not re.match(r"^[\d,\.\-]+$", remainder0):
            # code cell combined code+name ("1001000: เงินสด") and no other column had a name
            name = remainder0
        if dr_idx is not None and dr_idx < len(r) and cr_idx < len(r):
            d_val, c_val = r[dr_idx], r[cr_idx]
            dr = float(d_val) if isinstance(d_val, (int, float)) else 0.0
            cr = float(c_val) if isinstance(c_val, (int, float)) else 0.0
        else:
            # fallback heuristic: the last two numeric cells are the ending Dr / Cr pair
            nums = [c for c in r if isinstance(c, (int, float))]
            dr = cr = 0.0
            if len(nums) >= 2:
                dr, cr = float(nums[-2]), float(nums[-1])
            elif len(nums) == 1:
                dr = float(nums[0])
        out.append((code, name, dr, cr))

    imported = len(out)
    if imported == 0 and seen > 0:
        warnings.append(f"parsed 0 accounts from {seen} non-blank source row(s) below the header — this is "
                        "almost certainly an account-code detection failure, NOT a genuinely-empty TB; "
                        "inspect --client-tb's layout by hand before trusting an empty TB")
    elif imported < seen:
        warnings.append(f"saw {seen} non-blank source row(s) below the header but only {imported} yielded a "
                        f"leading numeric account code ({seen - imported} skipped) — verify none of those "
                        "skipped rows are real accounts")

    if not out:
        warnings.append("no account rows parsed from --client-tb; produced an empty TB to fill by hand")
        return out

    total_dr = sum(d for _, _, d, _ in out)
    total_cr = sum(c for _, _, _, c in out)
    imbalance = total_dr - total_cr
    balanced = abs(imbalance) < 0.01
    if closing:
        msg = (f"imported {imported} accounts from --client-tb using header-matched closing-balance columns "
              f"('{dr_hdr}' / '{cr_hdr}') — Σdebit={total_dr:,.2f} Σcredit={total_cr:,.2f}")
    else:
        msg = (f"imported {imported} accounts from --client-tb — no closing-balance header "
              "(ปลายงวด/ยกไป/คงเหลือ) found; fell back to the 'last two numeric cells' heuristic "
              f"(client exports vary; VERIFY the Dr/Cr columns landed correctly) — "
              f"Σdebit={total_dr:,.2f} Σcredit={total_cr:,.2f}")
    if balanced:
        msg += " — balances to zero"
    else:
        msg += (f" — DOES NOT BALANCE (Σdebit-Σcredit = {imbalance:,.2f}); the Dr/Cr columns are "
               "very likely misaligned or rows are missing — fix the TB by hand before relying on it")
    warnings.append(msg)
    return out


# ── styling helpers ────────────────────────────────────────────────────────────────
def F(**kw) -> Font:
    """A font in the locked FS typeface. Always name the font explicitly: a bare
    Font(bold=True) carries no name and Excel silently falls back to Calibri."""
    return Font(name=FS_FONT, size=FS_SIZE, **kw)


PLAIN = F()
BOLD = F(bold=True)
HEAD = F(bold=True)
WARN = F(color="9C6210")
THIN = Side(style="thin", color="BBBBBB")
RULE = Side(style="thin", color="000000")
TOPBORDER = Border(top=THIN)
TOPBOT = Border(top=THIN, bottom=Side(style="double", color="888888"))
HEADRULE = Border(bottom=RULE)   # the rule under the period line, across A:G
CENTER = Alignment(horizontal="center")


def money(ws, cell):
    ws[cell].number_format = NUMFMT


def sumif(cap: str, col: str) -> str:
    return f'=SUMIF({MAP}!$H:$H,"{cap}",{MAP}!${col}:${col})'


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("context", help="path to the client's CONTEXT.md")
    ap.add_argument("--client-tb", help="path to the client's TB/GL export .xlsx to pre-load accounts")
    ap.add_argument("--tb-sheet", help="sheet name inside --client-tb (default: first sheet)")
    args = ap.parse_args()

    ctx_path = Path(args.context)
    if not ctx_path.exists():
        fail(f"CONTEXT.md not found: {ctx_path}")
    profile, confidence = parse_context(ctx_path)
    warnings: list[str] = []

    # Refuse to render off a broken/unverified CONTEXT.md — same gate as
    # audit-cover-report's render_cover_report.py and audit-planning's render_planning.py.
    verdict = context_verdict(ctx_path)
    if verdict and re.search(r"REVIEW|AMBIGUOUS|JOB NOT IN DB", verdict, re.I):
        fail(f"CONTEXT §2 validation verdict is unresolved ({verdict!r}) — do NOT auto-generate; "
             "resolve the job-number cross-check before running this skill.")

    company_raw = clean(profile.get("company_legal_name")) or clean(profile.get("company_name"))
    tax_id = clean(profile.get("tax_id"))
    period_end_raw = clean(profile.get("period_end"))
    required = {"company_name": company_raw, "tax_id": tax_id, "period_end": period_end_raw}
    missing = [k for k, v in required.items() if not v]
    if missing:
        fail(f"unresolved required field(s) in CONTEXT (⚠ or missing): {missing}. Resolve before rendering.")

    entity = profile.get("entity_type", "")
    if "หจก" in entity or "ห้างหุ้นส่วน" in entity:
        warnings.append("entity_type looks like หจก. (partnership) — v1 scaffolds บจ. equity only; "
                        "the equity section needs the partnership variant (see contract §6)")

    company = company_raw
    p_label, p_tag = period_str(profile)
    cy = be_year(profile.get("period_end", "")) or "⚠"
    py = be_year(profile.get("prior_period_end", "")) or (cy - 1 if isinstance(cy, int) else "⚠")

    director = director_name(profile, confidence, warnings)

    tb_rows = import_client_tb(Path(args.client_tb), args.tb_sheet, warnings) if args.client_tb else []

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # workbook default typeface (contract §10.1) — cells we style carry FS_FONT explicitly
    wb._named_styles["Normal"].font = Font(name=DEFAULT_FONT, size=FS_SIZE)

    # ── hidden placeholder + caption vocabulary ────────────────────────────────────
    wb.create_sheet("0000").sheet_state = "veryHidden"
    cap_ws = wb.create_sheet(CAPREF)
    cap_ws.sheet_state = "hidden"
    cap_ws["A1"] = "controlled FS caption vocabulary — do not edit (docs/financials-contract.md §4)"
    for i, cap in enumerate(ALL_CAPTIONS, start=2):
        cap_ws[f"A{i}"] = cap
    cap_range = f"{CAPREF}!$A$2:$A${len(ALL_CAPTIONS) + 1}"

    # ── TB sheet ───────────────────────────────────────────────────────────────────
    tb = wb.create_sheet(TB)
    tb["A1"] = company
    tb["A2"] = f"งบทดลอง (ปรับปรุงแล้ว) ณ วันที่ {p_label}"
    hdr = ["เลขที่บัญชี", "ชื่อบัญชี", "ยอดยกมา เดบิต", "ยอดยกมา เครดิต",
           "ปรับปรุง เดบิต", "ปรับปรุง เครดิต", "คงเหลือ เดบิต", "คงเหลือ เครดิต"]
    for j, h in enumerate(hdr, start=1):
        c = tb.cell(row=3, column=j, value=h)
        c.font = BOLD
    r = 4
    for code, name, dr, cr in tb_rows:
        tb.cell(row=r, column=1, value=code)
        tb.cell(row=r, column=2, value=name)
        # ending balances go into opening (C/D); adjustments (E/F) start empty; G/H = C+E-D-F
        tb.cell(row=r, column=3, value=dr).number_format = NUMFMT
        tb.cell(row=r, column=4, value=cr).number_format = NUMFMT
        tb.cell(row=r, column=7, value=f"=C{r}+E{r}-D{r}-F{r}").number_format = NUMFMT
        tb.cell(row=r, column=8, value=f"=D{r}+F{r}-C{r}-E{r}").number_format = NUMFMT
        r += 1
    tb.column_dimensions["A"].width = 14
    tb.column_dimensions["B"].width = 34
    for col in "CDEFGH":
        tb.column_dimensions[col].width = 15

    # ── Mapping sheet ──────────────────────────────────────────────────────────────
    mp = wb.create_sheet(MAP)
    mp["A1"] = company
    mp["A2"] = f"กระดาษทำการจัดประเภทผังบัญชี (Mapping) ปี {cy} เทียบ {py}"
    mp["A3"] = "⚠ กรอกคอลัมน์ H (รายการในงบการเงิน) ให้ทุกบัญชี — เลือกจากรายการมาตรฐาน (dropdown)"
    mp["A3"].font = Font(color="9C6210")
    mhdr = ["ผังบัญชี", "ชื่อบัญชี", "ยอด เดบิต", "ยอด เครดิต",
            "ยอดปีก่อน เดบิต", "ยอดปีก่อน เครดิต", "หมวดตามหมายเหตุ", "รายการในงบการเงิน"]
    for j, h in enumerate(mhdr, start=1):
        mp.cell(row=5, column=j, value=h).font = BOLD
    r = 6
    for code, name, _dr, _cr in tb_rows:
        mp.cell(row=r, column=1, value=code)
        mp.cell(row=r, column=2, value=name)
        mp.cell(row=r, column=3, value=f'=SUMIF({TB}!$A:$A,A{r},{TB}!$G:$G)').number_format = NUMFMT
        mp.cell(row=r, column=4, value=f'=SUMIF({TB}!$A:$A,A{r},{TB}!$H:$H)').number_format = NUMFMT
        r += 1
    map_last = max(r - 1, 6)
    # dropdown on H restricting to the controlled vocabulary
    dv = DataValidation(type="list", formula1=f"={cap_range}", allow_blank=True)
    dv.error = "เลือกจากรายการมาตรฐาน (NPAE) เท่านั้น"
    dv.errorTitle = "caption ไม่ถูกต้อง"
    mp.add_data_validation(dv)
    dv.add(f"H6:H{map_last + 50}")
    mp.column_dimensions["A"].width = 14
    mp.column_dimensions["B"].width = 32
    for col in "CDEF":
        mp.column_dimensions[col].width = 14
    mp.column_dimensions["G"].width = 30
    mp.column_dimensions["H"].width = 32

    # ── งบการเงิน sheet: layout plan first, then render ─────────────────────────────
    # The sheet is a STACK OF PRINTED PAGES, not a continuous table (contract §10.3): each
    # statement is padded to a fixed ROWS_PER_PAGE grid and closed with an explicit page
    # break, so a statement never splits and every page carries its own header.
    # Build an ordered list of row specs; the row number is its position. Statement lines
    # reference note-detail rows, so we resolve positions in a plan pass then render.
    plan: list[dict] = []
    breaks: list[int] = []

    def add(kind, **kw) -> int:
        plan.append({"kind": kind, **kw})
        return len(plan)  # 1-based row number of the row just added

    def page_start() -> int:
        return breaks[-1] if breaks else 0

    def pad_to(offset: int) -> None:
        """Pad so the next added row lands at 1-based `offset` within the current page."""
        while len(plan) < page_start() + offset - 1:
            add("blank")

    def endpage(label: str) -> None:
        used = len(plan) - page_start()
        if used > ROWS_PER_PAGE:
            warnings.append(f"หน้า '{label}' ใช้ {used} แถว เกินกริด {ROWS_PER_PAGE} แถว/หน้า — "
                            "Excel จะตัดหน้าเอง ทำให้งบขาดกลางหน้า (ดู contract §10.3)")
        while len(plan) % ROWS_PER_PAGE:
            add("blank")
        breaks.append(len(plan))

    is_partnership = "หจก" in entity or "ห้างหุ้นส่วน" in entity
    period_line = f"สำหรับปีสิ้นสุดวันที่ {p_label}"

    def stmt_head(title, date_line, *, first=False, bare=False):
        """The page furniture every statement page opens with (contract §3.1).

        `bare` drops the unit + column header: the notes page is prose, and its note-detail
        area further down carries its own column header.
        """
        add("title", text=company if first else "=+A1", style="company")
        add("title", text=title, style="stmt")
        date_row = add("title", text=date_line, style="date")
        add("blank")
        if not bare:
            add("unit")
            add("colhdr", first=first)
        return date_row

    # --- page 1: balance sheet, assets side. Rows 1–3 are the literals every later page
    #     refers back to with '=+A1' / '=+A2' / '=+A3'.
    stmt_head("งบแสดงฐานะการเงิน", f"ณ วันที่ {p_label}", first=True)
    add("section", text="สินทรัพย์")
    add("blank")
    add("subhdr", text="สินทรัพย์หมุนเวียน")
    for cap, _ in CUR_ASSETS:
        add("cap", cap=cap)
    add("subtotal", text="รวมสินทรัพย์หมุนเวียน", group="cur_assets")
    add("blank")
    add("subhdr", text="สินทรัพย์ไม่หมุนเวียน")
    for cap, _ in NONCUR_ASSETS:
        add("cap", cap=cap)
    add("subtotal", text="รวมสินทรัพย์ไม่หมุนเวียน", group="noncur_assets")
    add("blank")
    add("total", text="รวมสินทรัพย์", group="assets", name="FS_TOTAL_ASSETS")
    # director sign-off block, anchored near the foot of the page like ground truth
    pad_to(27)
    if not is_partnership:
        add("sig_approval")
    add("blank")
    add("sig_certify")
    add("blank")
    add("sig_line")
    add("sig_name")
    endpage("งบแสดงฐานะการเงิน — สินทรัพย์")

    # --- page 2: balance sheet, liabilities + equity side ---
    stmt_head("=+A2", "=+A3")
    add("section", text="หนี้สินและส่วนของผู้ถือหุ้น")
    add("blank")
    add("subhdr", text="หนี้สินหมุนเวียน")
    for cap, _ in CUR_LIAB:
        add("cap", cap=cap)
    add("subtotal", text="รวมหนี้สินหมุนเวียน", group="cur_liab")
    add("blank")
    add("subhdr", text="หนี้สินไม่หมุนเวียน")
    for cap, _ in NONCUR_LIAB:
        add("cap", cap=cap)
    add("subtotal", text="รวมหนี้สินไม่หมุนเวียน", group="noncur_liab")
    add("total", text="รวมหนี้สิน", group="liab")
    add("blank")
    add("subhdr", text="ส่วนของผู้ถือหุ้น")
    add("cap", cap=SHARE_CAP[0])
    add("re_bs")  # retained earnings (from equity statement closing)
    add("subtotal", text="รวมส่วนของผู้ถือหุ้น", group="equity", name="FS_TOTAL_EQUITY")
    add("blank")
    add("total", text="รวมหนี้สินและส่วนของผู้ถือหุ้น", group="liabeq", name="FS_TOTAL_LIAB_EQUITY")
    add("checkrow")  # BS tie-out flag — rendered in column I, outside the print area
    endpage("งบแสดงฐานะการเงิน — หนี้สินและส่วนของผู้ถือหุ้น")

    # --- page 3: income statement ---
    is_date_row = stmt_head("งบกำไรขาดทุน", period_line)
    add("section", text="รายได้")
    for cap, _ in REVENUE:
        add("cap", cap=cap)
    add("subtotal", text="รวมรายได้", group="revenue")
    add("blank")
    add("section", text="ค่าใช้จ่าย")
    for cap, _ in EXPENSES:
        add("cap", cap=cap)
    add("subtotal", text="รวมค่าใช้จ่าย", group="expenses")
    add("blank")
    add("pbt", text="กำไร(ขาดทุน)ก่อนภาษีเงินได้", name="TAX_NET_PROFIT")
    add("taxline", text="ค่าใช้จ่ายภาษีเงินได้")
    add("netprofit", text="กำไร(ขาดทุน)สุทธิ", name="FS_NET_PROFIT")
    endpage("งบกำไรขาดทุน")

    # --- page 4: statement of changes in equity (minimal) ---
    stmt_head("งบแสดงการเปลี่ยนแปลงส่วนของผู้ถือหุ้น", f"=+A{is_date_row}")
    add("eq_open", text="ยอดคงเหลือต้นปี — กำไร(ขาดทุน)สะสม")
    add("eq_profit", text="กำไร(ขาดทุน)สุทธิสำหรับปี")
    add("eq_close", text="ยอดคงเหลือปลายปี — กำไร(ขาดทุน)สะสม")
    endpage("งบแสดงการเปลี่ยนแปลงส่วนของผู้ถือหุ้น")

    # --- page 5+: notes. Everything from the note-detail marker down is rebuilt by skill
    #     5.5 (expand_notes), so it is deliberately the LAST thing on the sheet and is left
    #     to flow across pages naturally rather than onto the fixed grid.
    stmt_head("หมายเหตุประกอบงบการเงิน", f"=+A{is_date_row}", bare=True)
    add("note_general")
    add("note_basis")
    add("note_policy")
    add("section", text="รายละเอียดประกอบ (Note detail — รายบรรทัดเพิ่มโดย audit-financials)")
    add("colhdr")
    for cap in ALL_CAPTIONS:
        add("note_detail", cap=cap)

    # resolve row numbers
    for i, spec in enumerate(plan, start=1):
        spec["row"] = i
    cap_note_row = {s["cap"]: s["row"] for s in plan if s["kind"] == "note_detail"}
    group_rows: dict[str, list[int]] = {}
    for s in plan:
        if s["kind"] == "cap":
            grp = _group_of(s["cap"])
            group_rows.setdefault(grp, []).append(s["row"])
    # named single rows resolved during render
    named: dict[str, int] = {}
    row_of = {"pbt": None, "netprofit": None, "eq_close": None, "re_bs": None,
              "eq_open": None, "eq_profit": None, "taxline": None}
    for s in plan:
        if s["kind"] in row_of:
            row_of[s["kind"]] = s["row"]

    ws = wb.create_sheet(FS)  # final sheet order is fixed at the end

    def rng(group):
        rows = sorted(group_rows.get(group, []))
        if not rows:
            return None
        return f"E{rows[0]}:E{rows[-1]}", f"G{rows[0]}:G{rows[-1]}"

    # composite subtotal groups (assets = cur+noncur, etc.)
    composite = {
        "assets": ["cur_assets", "noncur_assets"],
        "liab": ["cur_liab", "noncur_liab"],
        "liabeq": ["cur_liab", "noncur_liab", "equity_all"],
    }

    first_colhdr: int | None = None

    for s in plan:
        row = s["row"]
        k = s["kind"]
        a = ws.cell(row=row, column=1)
        a.font = PLAIN  # default; the bold kinds below override it
        if k == "title":
            # titles are merged across A:G and centred, with a rule under the period line
            a.value = s["text"]
            a.font = BOLD
            a.alignment = CENTER
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            if s.get("style") == "date":
                for c in range(1, 8):
                    ws.cell(row=row, column=c).border = HEADRULE
        elif k == "unit":
            u = ws.cell(row=row, column=5, value="หน่วย : บาท")
            u.font = BOLD
            u.alignment = CENTER
        elif k == "colhdr":
            # later pages reference the first column header so the years stay in one place
            src = first_colhdr if not s.get("first") and first_colhdr else None
            ws.cell(row=row, column=3, value="หมายเหตุ").font = BOLD
            e = ws.cell(row=row, column=5, value=f"=+E{src}" if src else cy)
            g = ws.cell(row=row, column=7, value=f"=+G{src}" if src else py)
            e.font = g.font = BOLD
            e.alignment = g.alignment = CENTER
            ws.cell(row=row, column=3).alignment = CENTER
            if first_colhdr is None:
                first_colhdr = row
        elif k == "sig_approval":
            a.value = ("งบการเงินนี้ได้รับอนุมัติจากที่ประชุมใหญ่สามัญผู้ถือหุ้น "
                       f"ครั้งที่ 1/{cy + 1 if isinstance(cy, int) else '⟨ปี⟩'} "
                       "เมื่อวันที่.......................................................")
            a.font = PLAIN
        elif k == "sig_certify":
            a.value = "ขอรับรองว่าถูกต้อง"
            a.font = PLAIN
        elif k == "sig_line":
            a.value = ("ลงชื่อ...................................................."
                       + ("หุ้นส่วนผู้จัดการ" if is_partnership else "กรรมการ"))
            a.font = PLAIN
        elif k == "sig_name":
            # never invent the name: unresolved/low-confidence → dotted blank for the human
            a.value = f"       ( {director} )" if director else "       ( .......................................... )"
            a.font = PLAIN
        elif k in ("section", "subhdr"):
            a.value = s["text"]
            a.font = BOLD
        elif k == "cap":
            cap = s["cap"]
            a.value = cap
            nr = cap_note_row[cap]
            nc = ws.cell(row=row, column=3, value=cap_note_index(cap))
            nc.alignment = CENTER
            nc.font = PLAIN
            _put(ws, f"E{row}", f"=+E{nr}")
            _put(ws, f"G{row}", f"=+G{nr}")
        elif k == "re_bs":
            a.value = RETAINED[0]
            _put(ws, f"E{row}", f"=+E{row_of['eq_close']}")
            _put(ws, f"G{row}", 0)
        elif k in ("subtotal", "total"):
            a.value = s["text"]
            a.font = BOLD
            groups = composite.get(s["group"], [s["group"]])
            # equity subtotal must include share cap row + re_bs row
            if s["group"] == "equity":
                erows = sorted(group_rows.get("equity", []) + [row_of["re_bs"]])
                _put(ws, f"E{row}", f"=SUM(E{erows[0]}:E{erows[-1]})")
                _put(ws, f"G{row}", f"=SUM(G{erows[0]}:G{erows[-1]})")
                group_rows["equity_all"] = erows
            else:
                parts_e, parts_g = [], []
                for g in groups:
                    rr = rng(g)
                    if rr:
                        parts_e.append(f"SUM({rr[0]})")
                        parts_g.append(f"SUM({rr[1]})")
                _put(ws, f"E{row}", "=" + "+".join(parts_e) if parts_e else "=0")
                _put(ws, f"G{row}", "=" + "+".join(parts_g) if parts_g else "=0")
            ws[f"E{row}"].border = TOPBORDER
            ws[f"G{row}"].border = TOPBORDER
            if s.get("name"):
                named[s["name"] + "_CY"] = row
                named[s["name"] + "_PY"] = row
        elif k == "pbt":
            a.value = s["text"]
            a.font = BOLD
            rev = rng("revenue")
            exp = rng("expenses")
            _put(ws, f"E{row}", f"=SUM({rev[0]})-SUM({exp[0]})")
            _put(ws, f"G{row}", f"=SUM({rev[1]})-SUM({exp[1]})")
            named["TAX_NET_PROFIT"] = row
        elif k == "taxline":
            a.value = s["text"]
            _put(ws, f"E{row}", "=TAX_EXPENSE")  # defined name → ภาษีเงินได้ computed cell
            _put(ws, f"G{row}", 0)
        elif k == "netprofit":
            a.value = s["text"]
            a.font = BOLD
            _put(ws, f"E{row}", f"=E{row_of['pbt']}-E{row_of['taxline']}")
            _put(ws, f"G{row}", f"=G{row_of['pbt']}-G{row_of['taxline']}")
            ws[f"E{row}"].border = TOPBOT
            ws[f"G{row}"].border = TOPBOT
            named["FS_NET_PROFIT_CY"] = row
            named["FS_NET_PROFIT_PY"] = row
        elif k == "eq_open":
            a.value = s["text"]
            _put(ws, f"E{row}", sumif(RETAINED[0], "D"))  # opening RE = mapped RE-account balance
        elif k == "eq_profit":
            a.value = s["text"]
            _put(ws, f"E{row}", f"=+E{row_of['netprofit']}")
        elif k == "eq_close":
            a.value = s["text"]
            a.font = BOLD
            _put(ws, f"E{row}", f"=E{row_of['eq_open']}+E{row_of['eq_profit']}")
            ws[f"E{row}"].border = TOPBORDER
        elif k == "checkrow":
            # tie-out lives in column I — OUTSIDE the $A:$G print area (contract §3.1/§10.2),
            # so it never prints on the client's งบ but stays live for the auditor and 5.5.
            lbl = ws.cell(row=row, column=CHECK_COL,
                          value="ตรวจสอบ: สินทรัพย์ = หนี้สิน+ทุน (ต้องเป็น 0)")
            lbl.font = Font(name=FS_FONT, size=FS_SIZE, italic=True, color="9C6210")
            ta = named.get("FS_TOTAL_ASSETS_CY")
            tle = named.get("FS_TOTAL_LIAB_EQUITY_CY")
            if ta and tle:
                chk = ws.cell(row=row, column=CHECK_COL + 1, value=f"=E{ta}-E{tle}")
                chk.number_format = NUMFMT
        elif k == "note_general":
            a.value = "1. ข้อมูลทั่วไป"
            a.font = BOLD
            ws.cell(row=row, column=1)  # detail appended on next rows via note_basis/policy blocks
        elif k == "note_basis":
            a.value = "2. เกณฑ์การจัดทำงบการเงิน"
            a.font = BOLD
        elif k == "note_policy":
            a.value = "3. นโยบายการบัญชีที่สำคัญ"
            a.font = BOLD
        elif k == "note_detail":
            cap = s["cap"]
            a.value = cap
            a.alignment = Alignment(indent=1)
            nc = ws.cell(row=row, column=3, value=cap_note_index(cap))
            nc.alignment = CENTER
            nc.font = PLAIN
            side = SIDE[cap]
            _put(ws, f"E{row}", sumif(cap, side))
            _put(ws, f"G{row}", sumif(cap, "E" if side == "C" else "F"))
        elif k == "blank":
            pass
        # apply money format + the FS font to E/G on value rows
        if k in ("cap", "subtotal", "total", "pbt", "taxline", "netprofit", "re_bs",
                 "eq_open", "eq_profit", "eq_close", "note_detail"):
            for col in ("E", "G"):
                money(ws, f"{col}{row}")
                if ws[f"{col}{row}"].font.name != FS_FONT:
                    ws[f"{col}{row}"].font = PLAIN

    # ── print setup for the งบ (contract §10) — this sheet is a client deliverable ──────
    for col, w in COL_W.items():
        ws.column_dimensions[col].width = w
    last_row = len(plan)
    for r in range(1, last_row + 1):
        ws.row_dimensions[r].height = ROW_H

    ws.print_area = f"$A$1:$G${last_row}"
    ps = ws.page_setup
    ps.paperSize = ws.PAPERSIZE_A4
    ps.orientation = ws.ORIENTATION_PORTRAIT
    ps.fitToWidth = 1
    ps.fitToHeight = 0          # 1 page wide, flow as tall as it needs
    ps.firstPageNumber = FIRST_PAGE_NUMBER   # bound behind ใบปะหน้า + หน้ารายงานผู้สอบ
    ps.useFirstPageNumber = True
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.79, right=0.2, top=0.59, bottom=0.39,
                                  header=0.51, footer=0.28)
    ws.oddFooter.right.text = "&P"
    ws.oddFooter.right.font = f"{FOOTER_FONT},Regular"
    ws.oddFooter.right.size = 14
    # explicit breaks so each statement owns its page and never splits mid-statement
    for b in breaks:
        ws.row_breaks.append(Break(id=b))

    # ── ภาษีเงินได้ (tax shell) ─────────────────────────────────────────────────────
    tx = wb.create_sheet(TAX)
    tx["A1"] = company
    tx["A2"] = f"การคำนวณภาษีเงินได้นิติบุคคล ปี {cy}"
    tx["A2"].font = BOLD
    trows = [
        ("กำไร(ขาดทุน)สุทธิทางบัญชีก่อนภาษี", f"=+'{FS}'!E{row_of['pbt']}", None),
        ("บวก รายจ่ายต้องห้าม", "⚠ ⟨human⟩", "warn"),
        ("หัก รายได้ที่ได้รับยกเว้น", "⚠ ⟨human⟩", "warn"),
        ("กำไร(ขาดทุน)สุทธิทางภาษี", None, "nettax"),
        ("หัก ยกเว้น SME 300,000 บาทแรก", "⚠ ⟨human: ถ้าเข้าเงื่อนไข SME⟩", "warn"),
        ("ฐานภาษี", None, "base"),
        ("อัตราภาษี", 0.15, "rate"),
        ("ภาษีเงินได้คำนวณ", None, "computed"),
        ("หัก ภาษีหัก ณ ที่จ่าย / ภงด.51", "⚠ ⟨human⟩", "warn"),
        ("ภาษีเงินได้ที่ต้องชำระ(ชำระเกิน)", None, "payable"),
    ]
    tr = {}
    r = 4
    for label, val, key in trows:
        tx.cell(row=r, column=1, value=label)
        cell = tx.cell(row=r, column=3)
        if isinstance(val, (int, float)):
            cell.value = val
        elif isinstance(val, str):
            cell.value = val
        if key:
            tr[key] = r
        r += 1
    # computed formulas (best-effort; add-backs are human, so N() treats blanks as 0)
    tx.cell(row=tr["nettax"], column=3,
            value=f"=C4+N(C{tr['nettax']-2})-N(C{tr['nettax']-1})").number_format = NUMFMT
    tx.cell(row=tr["base"], column=3,
            value=f"=MAX(0,C{tr['nettax']}-N(C{tr['base']-1}))").number_format = NUMFMT
    tx.cell(row=tr["computed"], column=3,
            value=f"=ROUND(C{tr['base']}*C{tr['rate']},2)").number_format = NUMFMT
    tx.cell(row=tr["payable"], column=3,
            value=f"=C{tr['computed']}-N(C{tr['payable']-1})").number_format = NUMFMT
    tx["C3"].number_format = NUMFMT
    tx.cell(row=tr["rate"], column=3).number_format = "0%"
    tx.column_dimensions["A"].width = 40
    tx.column_dimensions["C"].width = 18

    # tax expense on IS references the computed tax
    wb.defined_names["TAX_EXPENSE"] = DefinedName("TAX_EXPENSE", attr_text=f"'{TAX}'!$C${tr['computed']}")
    wb.defined_names["TAX_NET_PROFIT"] = DefinedName("TAX_NET_PROFIT", attr_text=f"'{FS}'!$E${row_of['pbt']}")
    wb.defined_names["TAX_PAYABLE"] = DefinedName("TAX_PAYABLE", attr_text=f"'{TAX}'!$C${tr['payable']}")

    # ── ปรับปรุง (AJE grid) ─────────────────────────────────────────────────────────
    aj = wb.create_sheet(AJE)
    aj["A1"] = company
    aj["A2"] = f"รายการปรับปรุง (Adjusting entries) ปี {cy}"
    aj["A2"].font = BOLD
    ajhdr = ["เลขที่", "คำอธิบายรายการ", "Ref", "เดบิต (งบดุล)", "เครดิต (งบดุล)",
             "เดบิต (กำไรขาดทุน)", "เครดิต (กำไรขาดทุน)"]
    for j, h in enumerate(ajhdr, start=1):
        aj.cell(row=3, column=j, value=h).font = BOLD
    aj["A4"] = "⚠ ⟨human: ลงรายการปรับปรุงตามดุลพินิจ — สคริปต์ไม่สร้างรายการเอง⟩"
    aj["A4"].font = Font(color="9C6210")
    aj.column_dimensions["A"].width = 8
    aj.column_dimensions["B"].width = 40
    for col in "DEFG":
        aj.column_dimensions[col].width = 15

    # ── defined names for FS totals ────────────────────────────────────────────────
    for nm, rownum in named.items():
        col = "E" if nm.endswith("_CY") else ("G" if nm.endswith("_PY") else "E")
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"'{FS}'!${col}${rownum}")

    # order sheets: 0000, งบการเงิน, TB, Mapping, ปรับปรุง, ภาษีเงินได้, _captions
    order = ["0000", FS, TB, MAP, AJE, TAX, CAPREF]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

    # ── write ──────────────────────────────────────────────────────────────────────
    out_dir = ctx_path.parent / "WP"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"4 งบการเงิน {company} {p_tag}.xlsx"
    wb.save(out_path)

    if not tb_rows:
        warnings.append("no client TB imported — TB and Mapping are empty layouts; paste the client's "
                        "trial balance into TB and the accounts into Mapping (or re-run with --client-tb)")
    warnings.append("Mapping!H (การจัดประเภทแต่ละบัญชีเข้า caption งบ) is a human judgment call — the งบ "
                    "stays zero until it is filled from the dropdown")
    warnings.append("v1 covers บจ. going-concern only; หจก./งบเลิก/ปีแรก need their variants (contract §6)")

    print(json.dumps({"ok": True, "output": str(out_path), "accounts_imported": len(tb_rows),
                      "warnings": warnings}, ensure_ascii=False))


# ── caption grouping / note numbering ───────────────────────────────────────────────
_GROUP = {}
for _c, _ in CUR_ASSETS:
    _GROUP[_c] = "cur_assets"
for _c, _ in NONCUR_ASSETS:
    _GROUP[_c] = "noncur_assets"
for _c, _ in CUR_LIAB:
    _GROUP[_c] = "cur_liab"
for _c, _ in NONCUR_LIAB:
    _GROUP[_c] = "noncur_liab"
_GROUP[SHARE_CAP[0]] = "equity"
_GROUP[RETAINED[0]] = "equity"
for _c, _ in REVENUE:
    _GROUP[_c] = "revenue"
for _c, _ in EXPENSES:
    _GROUP[_c] = "expenses"
_NOTE_INDEX = {cap: i + 4 for i, cap in enumerate(ALL_CAPTIONS)}  # notes start at 4


def _group_of(cap):
    return _GROUP[cap]


def cap_note_index(cap):
    return _NOTE_INDEX.get(cap, "")


def _put(ws, cell, value):
    ws[cell] = value


if __name__ == "__main__":
    main()
